"""XLSX document renderer."""

from __future__ import annotations

from io import BytesIO

from lewlm.core.errors import DocumentGenerationError
from lewlm.documents.ir.models import (
    CalloutBlock,
    DocumentIR,
    DocumentOutputFormat,
    ImageBlock,
    ListBlock,
    ParagraphBlock,
    TableBlock,
)
from lewlm.documents.render.base import DocumentRenderer
from lewlm.documents.validators.ir import DocumentIRValidator


class XlsxDocumentRenderer(DocumentRenderer):
    output_format = DocumentOutputFormat.XLSX
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    file_extension = ".xlsx"

    def __init__(self, validator: DocumentIRValidator) -> None:
        self.validator = validator

    def render(self, document: DocumentIR) -> bytes:
        self.validator.validate(document)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError as exc:
            raise DocumentGenerationError(
                "XLSX generation requires the `openpyxl` dependency.",
                details={"renderer": self.output_format.value},
            ) from exc

        workbook = Workbook()
        first_sheet = workbook.active
        for index, section in enumerate(document.sections):
            worksheet = first_sheet if index == 0 else workbook.create_sheet()
            worksheet.title = self._sheet_title(section.heading or f"Section {index + 1}")
            row = 1
            if index == 0:
                worksheet.cell(row=row, column=1, value=document.title)
                worksheet.cell(row=row, column=1).font = Font(bold=True)
                row += 2
            if section.heading:
                worksheet.cell(row=row, column=1, value=section.heading)
                worksheet.cell(row=row, column=1).font = Font(bold=True)
                row += 2
            for block in section.blocks:
                row = self._render_block(worksheet, block, row)
                row += 1

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _render_block(self, worksheet, block, row: int) -> int:
        if isinstance(block, ParagraphBlock):
            worksheet.cell(row=row, column=1, value=block.text)
            return row
        if isinstance(block, ListBlock):
            for index, item in enumerate(block.items, start=1):
                prefix = f"{index}. " if block.ordered else "- "
                worksheet.cell(row=row, column=1, value=f"{prefix}{item}")
                row += 1
            return row - 1
        if isinstance(block, CalloutBlock):
            prefix = block.kind.upper()
            if block.title:
                prefix = f"{prefix}: {block.title}"
            worksheet.cell(row=row, column=1, value=f"{prefix} {block.body}")
            return row
        if isinstance(block, ImageBlock):
            worksheet.cell(row=row, column=1, value=f"[Image] {block.alt_text}")
            if block.caption:
                worksheet.cell(row=row + 1, column=1, value=block.caption)
                return row + 1
            return row
        if isinstance(block, TableBlock):
            headers = block.headers or [f"column_{index + 1}" for index in range(len(block.rows[0]))]
            if block.caption:
                worksheet.cell(row=row, column=1, value=block.caption)
                row += 1
            for column, header in enumerate(headers, start=1):
                worksheet.cell(row=row, column=column, value=header)
            row += 1
            for values in block.rows:
                for column, value in enumerate(values, start=1):
                    worksheet.cell(row=row, column=column, value=value)
                row += 1
            return row - 1
        return row

    def _sheet_title(self, value: str) -> str:
        sanitized = "".join(char for char in value if char not in {"\\", "/", "*", "[", "]", ":", "?"}).strip()
        return (sanitized or "Sheet")[:31]

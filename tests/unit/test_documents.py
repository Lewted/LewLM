from __future__ import annotations

import json
from io import BytesIO

import pytest

pytest.importorskip("openpyxl")
pytest.importorskip("docx")

from openpyxl import load_workbook

from lewlm.core.errors import DocumentValidationError
from lewlm.documents.ir.models import DocumentIR, DocumentOutputFormat, DocumentSection, TableBlock
from lewlm.documents.service import DocumentGenerationService
from lewlm.documents.skills.models import DOCUMENT_TRANSFORM_REQUEST_ADAPTER
from lewlm.documents.skills.service import DocumentTransformService


def test_document_generation_service_renders_supported_formats(sample_document_ir: DocumentIR) -> None:
    service = DocumentGenerationService()

    text_artifact = service.generate(sample_document_ir, output_format=DocumentOutputFormat.TEXT)
    markdown_artifact = service.generate(sample_document_ir, output_format=DocumentOutputFormat.MARKDOWN)
    json_artifact = service.generate(sample_document_ir, output_format=DocumentOutputFormat.JSON)
    csv_artifact = service.generate(sample_document_ir, output_format=DocumentOutputFormat.CSV)
    docx_artifact = service.generate(sample_document_ir, output_format=DocumentOutputFormat.DOCX)
    pdf_artifact = service.generate(sample_document_ir, output_format=DocumentOutputFormat.PDF)
    xlsx_artifact = service.generate(sample_document_ir, output_format=DocumentOutputFormat.XLSX)

    assert text_artifact.content.decode("utf-8").startswith("Quarterly Operations Summary")
    assert "# Quarterly Operations Summary" in markdown_artifact.content.decode("utf-8")
    assert json.loads(json_artifact.content.decode("utf-8"))["title"] == "Quarterly Operations Summary"
    assert csv_artifact.content.decode("utf-8").splitlines()[0] == "Category,Amount"
    assert docx_artifact.content.startswith(b"PK")
    assert pdf_artifact.content.startswith(b"%PDF")
    workbook = load_workbook(BytesIO(xlsx_artifact.content))
    assert workbook.sheetnames == ["Summary", "Budget"]
    assert workbook["Budget"]["A4"].value == "Category"
    assert workbook["Budget"]["B5"].value == "1200"


def test_document_validation_rejects_inconsistent_table_rows() -> None:
    service = DocumentGenerationService()
    invalid_document = DocumentIR(
        title="Broken Table",
        sections=[
            DocumentSection(
                heading="Data",
                blocks=[TableBlock(headers=["A", "B"], rows=[["1", "2"], ["3"]])],
            ),
        ],
    )

    try:
        service.generate(invalid_document, output_format=DocumentOutputFormat.CSV)
    except DocumentValidationError as exc:
        assert exc.code == "document_validation_error"
    else:
        raise AssertionError("Expected DocumentValidationError for inconsistent table rows.")


def test_document_transform_service_supports_contract_and_receipt_skills(
    contract_transform_payload: dict[str, object],
    receipt_transform_payload: dict[str, object],
    branded_document_template_payload: dict[str, object],
    ocr_assisted_extraction_payload: dict[str, object],
    file_template_transform_payload: dict[str, object],
    document_compare_transform_payload: dict[str, object],
    meeting_transcript_notes_payload: dict[str, object],
    long_document_memo_payload: dict[str, object],
    speech_transcript_cleanup_payload: dict[str, object],
) -> None:
    service = DocumentTransformService()

    contract_request = DOCUMENT_TRANSFORM_REQUEST_ADAPTER.validate_python(contract_transform_payload)
    receipt_request = DOCUMENT_TRANSFORM_REQUEST_ADAPTER.validate_python(receipt_transform_payload)
    branded_request = DOCUMENT_TRANSFORM_REQUEST_ADAPTER.validate_python(branded_document_template_payload)
    ocr_request = DOCUMENT_TRANSFORM_REQUEST_ADAPTER.validate_python(ocr_assisted_extraction_payload)
    template_request = DOCUMENT_TRANSFORM_REQUEST_ADAPTER.validate_python(file_template_transform_payload)
    compare_request = DOCUMENT_TRANSFORM_REQUEST_ADAPTER.validate_python(document_compare_transform_payload)
    meeting_request = DOCUMENT_TRANSFORM_REQUEST_ADAPTER.validate_python(meeting_transcript_notes_payload)
    memo_request = DOCUMENT_TRANSFORM_REQUEST_ADAPTER.validate_python(long_document_memo_payload)
    cleanup_request = DOCUMENT_TRANSFORM_REQUEST_ADAPTER.validate_python(speech_transcript_cleanup_payload)

    contract_artifact = service.transform(contract_request)
    receipt_artifact = service.transform(receipt_request)
    branded_artifact = service.transform(branded_request)
    ocr_artifact = service.transform(ocr_request)
    template_artifact = service.transform(template_request)
    compare_artifact = service.transform(compare_request)
    meeting_artifact = service.transform(meeting_request)
    memo_artifact = service.transform(memo_request)
    cleanup_artifact = service.transform(cleanup_request)

    assert contract_artifact.content.startswith(b"PK")
    assert receipt_artifact.content.decode("utf-8").startswith("Description,Quantity,Unit Price,Line Total")
    branded_document = json.loads(branded_artifact.content.decode("utf-8"))
    assert branded_document["title"] == "LewLM Product Launch Brief"
    assert branded_document["header"]["left"] == "LewLM"
    assert branded_document["footer"]["center"] == "Internal planning use only"
    image_blocks = [
        block
        for section in branded_document["sections"]
        for block in section["blocks"]
        if block["type"] == "image"
    ]
    assert len(image_blocks) == 2
    assert any(block["role"] == "logo" for block in image_blocks)
    assert any(block["role"] == "image" for block in image_blocks)
    ocr_markdown = ocr_artifact.content.decode("utf-8")
    assert ocr_markdown.startswith("# Scanned Invoice Extraction")
    assert "## Extracted Fields" in ocr_markdown
    assert "Invoice Number | INV-2048 | extracted" in ocr_markdown
    assert "Payment Terms |  | missing" in ocr_markdown
    assert "Urgent handling for April restock." in ocr_markdown
    workbook = load_workbook(BytesIO(template_artifact.content))
    assert workbook["Overview"]["A1"].value == "Engagement Summary for Acme Corp"
    assert workbook["Overview"]["B9"].value == "On Track"
    comparison_workbook = load_workbook(BytesIO(compare_artifact.content))
    assert comparison_workbook["Overview"]["A7"].value == "Metric"
    assert comparison_workbook["Overview"]["B8"].value == "Baseline Agreement"
    assert comparison_workbook["Shared Segments"]["A3"].value == "- Shared scope line."
    meeting_markdown = meeting_artifact.content.decode("utf-8")
    assert meeting_markdown.startswith("# Project Kickoff Notes")
    assert "## Action Items" in meeting_markdown
    assert "Maya | Send revised delivery timeline | 2026-04-20" in meeting_markdown
    memo_markdown = memo_artifact.content.decode("utf-8")
    assert memo_markdown.startswith("# Platform Readiness Memo")
    assert "## Memo Summary" in memo_markdown
    assert "## Open Questions" in memo_markdown
    assert "What operator guidance is still missing for Linux and Windows rollouts?" in memo_markdown
    cleanup_markdown = cleanup_artifact.content.decode("utf-8")
    assert cleanup_markdown.startswith("# Customer Call Cleanup")
    assert "## Cleaned Transcript" in cleanup_markdown
    assert "Agent | Thanks everyone for joining today." in cleanup_markdown
    assert "Unknown | Follow up with legal for the final notice." in cleanup_markdown

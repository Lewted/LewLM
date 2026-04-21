from __future__ import annotations

import base64
import json
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from lewlm.cli.main import main


def test_end_to_end_http_chat_and_response_workflow(app_with_fake_runtime) -> None:
    with TestClient(app_with_fake_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        assert scan_response.status_code == 200

        manifests = scan_response.json()["manifests"]
        gguf_model_id = next(manifest["model_id"] for manifest in manifests if manifest["format_type"] == "gguf")

        chat_response = client.post(
            "/v1/chat/completions",
            json={
                "model": gguf_model_id,
                "messages": [{"role": "user", "content": "Summarize the plan"}],
            },
        )
        assert chat_response.status_code == 200
        assert chat_response.json()["choices"][0]["message"]["content"] == "Echo: Summarize the plan"

        response_api = client.post(
            "/v1/responses",
            json={"model": gguf_model_id, "input": "Turn this into notes"},
        )
        assert response_api.status_code == 200
        assert response_api.json()["output_text"] == "Echo: Turn this into notes"

        with client.stream(
            "POST",
            "/v1/chat/completions",
            json={
                "model": gguf_model_id,
                "messages": [{"role": "user", "content": "Stream this back"}],
                "stream": True,
            },
        ) as stream_response:
            assert stream_response.status_code == 200
            data_lines = [
                line.removeprefix("data: ")
                for line in stream_response.iter_lines()
                if line and line.startswith("data: ")
            ]

    assert data_lines[-1] == "[DONE]"
    stream_chunks = [json.loads(line) for line in data_lines[:-1]]
    streamed_text = "".join(
        chunk["choices"][0]["delta"].get("content", "")
        for chunk in stream_chunks
        if chunk["choices"][0]["delta"].get("content")
    )
    assert streamed_text == "Echo: Stream this back"


def test_end_to_end_model_lifecycle_and_event_workflow(app_with_fake_runtime) -> None:
    with TestClient(app_with_fake_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        manifests = scan_response.json()["manifests"]
        gguf_model_id = next(manifest["model_id"] for manifest in manifests if manifest["format_type"] == "gguf")

        warm_response = client.post(f"/v1/models/{gguf_model_id}/warm")
        assert warm_response.status_code == 200
        assert warm_response.json()["status"] == "warmed"

        with client.websocket_connect("/v1/events") as websocket:
            chat_response = client.post(
                "/v1/chat/completions",
                json={
                    "model": gguf_model_id,
                    "messages": [{"role": "user", "content": "Emit lifecycle events"}],
                },
            )
            assert chat_response.status_code == 200
            events = [websocket.receive_json() for _ in range(9)]

        unload_response = client.post(f"/v1/models/{gguf_model_id}/unload")
        assert unload_response.status_code == 200
        assert unload_response.json()["status"] == "unloaded"

    assert [event["type"] for event in events] == [
        "request.accepted",
        "operation.progress",
        "model.loading",
        "model.loaded",
        "operation.progress",
        "prefill.started",
        "operation.progress",
        "operation.progress",
        "request.completed",
    ]
    assert events[1]["payload"]["stage"] == "prompt_compiled"
    assert events[4]["payload"]["stage"] == "model_ready"
    assert events[6]["payload"]["reasoning_exposed"] is False


def test_end_to_end_document_ingest_event_workflow(app_with_fake_runtime, sample_ingest_sources) -> None:
    with TestClient(app_with_fake_runtime) as client:
        with client.websocket_connect("/v1/events") as websocket:
            ingest_response = client.post("/v1/documents/ingest", json={"paths": [str(sample_ingest_sources["pdf"])]})
            assert ingest_response.status_code == 200
            request_id = ingest_response.json()["request_id"]
            events = [websocket.receive_json() for _ in range(3)]

    assert [event["type"] for event in events] == [
        "document.parse.started",
        "operation.progress",
        "document.parse.completed",
    ]
    assert all(event["payload"]["request_id"] == request_id for event in events)


def test_end_to_end_http_document_ingest_workflow(app_with_fake_runtime, sample_ingest_sources) -> None:
    with TestClient(app_with_fake_runtime) as client:
        pdf_response = client.post("/v1/documents/ingest", json={"paths": [str(sample_ingest_sources["pdf"])]})
        image_response = client.post("/v1/documents/ingest", json={"paths": [str(sample_ingest_sources["image_bundle"])]})

    assert pdf_response.status_code == 200
    assert pdf_response.json()["request_id"]
    assert pdf_response.json()["document"]["title"] == "sample"
    assert pdf_response.json()["sources"][0]["source_type"] == "pdf"
    assert image_response.status_code == 200
    assert image_response.json()["request_id"]
    assert image_response.json()["sources"][0]["metadata"]["image_count"] == 2


def test_end_to_end_http_multimodal_workflow(
    app_with_fake_multimodal_runtime,
    sample_audio_bytes: bytes,
) -> None:
    with TestClient(app_with_fake_multimodal_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        manifests = scan_response.json()["manifests"]
        embedding_model_id = next(
            manifest["model_id"]
            for manifest in manifests
            if manifest["display_name"] == "e5-small-embed-mlx"
        )
        rerank_model_id = next(
            manifest["model_id"]
            for manifest in manifests
            if manifest["display_name"] == "bge-reranker-base-mlx"
        )
        audio_model_id = next(
            manifest["model_id"]
            for manifest in manifests
            if manifest["display_name"] == "whisper-mini-audio"
        )

        embeddings_response = client.post(
            "/v1/embeddings",
            json={"model": embedding_model_id, "input": "Plan the next milestone"},
        )
        rerank_response = client.post(
            "/v1/rerank",
            json={
                "model": rerank_model_id,
                "query": "milestone plan",
                "documents": ["milestone plan summary", "audio transcription", "routing diagnostics"],
            },
        )
        transcription_response = client.post(
            "/v1/audio/transcriptions",
            json={
                "model": audio_model_id,
                "audio_base64": base64.b64encode(sample_audio_bytes).decode("ascii"),
                "file_name": "workflow.wav",
            },
        )
        speech_response = client.post(
            "/v1/audio/speech",
            json={"model": audio_model_id, "input": "Milestone seven is underway.", "format": "wav"},
        )

    assert embeddings_response.status_code == 200
    assert embeddings_response.json()["model"] == embedding_model_id
    assert len(embeddings_response.json()["data"]) == 1

    assert rerank_response.status_code == 200
    assert rerank_response.json()["results"][0]["index"] == 0

    assert transcription_response.status_code == 200
    assert transcription_response.json()["request_id"]
    assert "workflow.wav" in transcription_response.json()["text"]

    assert speech_response.status_code == 200
    assert speech_response.json()["request_id"]
    assert base64.b64decode(speech_response.json()["audio_base64"]).startswith(b"RIFF")


def test_end_to_end_http_audio_event_workflow(
    app_with_fake_multimodal_runtime,
    sample_audio_bytes: bytes,
) -> None:
    with TestClient(app_with_fake_multimodal_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        audio_model_id = next(
            manifest["model_id"]
            for manifest in scan_response.json()["manifests"]
            if manifest["display_name"] == "whisper-mini-audio"
        )
        with client.websocket_connect("/v1/events") as websocket:
            transcription_response = client.post(
                "/v1/audio/transcriptions",
                json={
                    "model": audio_model_id,
                    "audio_base64": base64.b64encode(sample_audio_bytes).decode("ascii"),
                    "file_name": "workflow.wav",
                },
            )
            assert transcription_response.status_code == 200
            request_id = transcription_response.json()["request_id"]
            events = [websocket.receive_json() for _ in range(8)]

    assert [event["type"] for event in events] == [
        "request.accepted",
        "audio.transcription.started",
        "operation.progress",
        "model.loading",
        "model.loaded",
        "operation.progress",
        "audio.transcription.completed",
        "request.completed",
    ]
    assert all(event["payload"]["request_id"] == request_id for event in events)


def test_end_to_end_audio_workflow_emits_chunk_events_for_long_audio(
    app_with_fake_multimodal_runtime,
    long_sample_audio_bytes: bytes,
) -> None:
    with TestClient(app_with_fake_multimodal_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        audio_model_id = next(
            manifest["model_id"]
            for manifest in scan_response.json()["manifests"]
            if manifest["display_name"] == "whisper-mini-audio"
        )
        with client.websocket_connect("/v1/events") as websocket:
            transcription_response = client.post(
                "/v1/audio/transcriptions",
                json={
                    "model": audio_model_id,
                    "audio_base64": base64.b64encode(long_sample_audio_bytes).decode("ascii"),
                    "file_name": "workflow-long.wav",
                },
            )
            assert transcription_response.status_code == 200
            request_id = transcription_response.json()["request_id"]
            events = [websocket.receive_json() for _ in range(14)]

    assert [event["type"] for event in events] == [
        "request.accepted",
        "audio.transcription.started",
        "operation.progress",
        "model.loading",
        "model.loaded",
        "operation.progress",
        "audio.chunk",
        "operation.progress",
        "operation.progress",
        "audio.chunk",
        "operation.progress",
        "operation.progress",
        "audio.transcription.completed",
        "request.completed",
    ]
    assert all(event["payload"]["request_id"] == request_id for event in events)
    assert events[2]["payload"]["stage"] == "chunks_planned"
    assert [event["payload"]["chunk_index"] for event in events if event["type"] == "audio.chunk"] == [1, 2]
    assert [event["payload"]["stage"] for event in events if event["type"] == "operation.progress"] == [
        "chunks_planned",
        "chunk_started",
        "chunk_processed",
        "chunk_started",
        "chunk_processed",
        "segments_ready",
    ]


def test_end_to_end_streaming_chat_exposes_model_emitted_reasoning_in_final_chunk(app_with_fake_runtime) -> None:
    with TestClient(app_with_fake_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        gguf_model_id = next(manifest["model_id"] for manifest in scan_response.json()["manifests"] if manifest["format_type"] == "gguf")

        with client.stream(
            "POST",
            "/v1/chat/completions",
            json={
                "model": gguf_model_id,
                "messages": [{"role": "user", "content": "[emit-reasoning] Stream the answer"}],
                "reasoning_visibility": "raw_model_emitted",
                "stream": True,
            },
        ) as stream_response:
            assert stream_response.status_code == 200
            data_lines = [
                line.removeprefix("data: ")
                for line in stream_response.iter_lines()
                if line and line.startswith("data: ")
            ]

    assert data_lines[-1] == "[DONE]"
    stream_chunks = [json.loads(line) for line in data_lines[:-1]]
    streamed_text = "".join(
        chunk["choices"][0]["delta"].get("content", "")
        for chunk in stream_chunks
        if chunk["choices"][0]["delta"].get("content")
    )
    streamed_reasoning = "".join(
        chunk["choices"][0]["delta"]["reasoning"]["content"]
        for chunk in stream_chunks[:-1]
        if chunk["choices"][0]["delta"].get("reasoning")
    )
    assert streamed_text == "Echo: Stream the answer"
    assert streamed_reasoning == "Inspect the prompt before replying."
    assert stream_chunks[-1]["choices"][0]["delta"]["reasoning"] == {
        "visibility": "raw_model_emitted",
        "available": True,
        "content": "Inspect the prompt before replying.",
        "summary": None,
    }


def test_end_to_end_streaming_responses_expose_model_emitted_reasoning_chunks(app_with_fake_runtime) -> None:
    with TestClient(app_with_fake_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        gguf_model_id = next(manifest["model_id"] for manifest in scan_response.json()["manifests"] if manifest["format_type"] == "gguf")

        with client.stream(
            "POST",
            "/v1/responses",
            json={
                "model": gguf_model_id,
                "input": "[emit-reasoning] Stream the response",
                "reasoning_visibility": "raw_model_emitted",
                "stream": True,
            },
        ) as stream_response:
            assert stream_response.status_code == 200
            data_lines = [
                line.removeprefix("data: ")
                for line in stream_response.iter_lines()
                if line and line.startswith("data: ")
            ]

    assert data_lines[-1] == "[DONE]"
    stream_chunks = [json.loads(line) for line in data_lines[:-1]]
    streamed_text = "".join(chunk.get("delta", "") for chunk in stream_chunks if chunk.get("delta"))
    streamed_reasoning = "".join(
        chunk["reasoning"]["content"]
        for chunk in stream_chunks[:-1]
        if chunk.get("reasoning")
    )
    assert streamed_text == "Echo: Stream the response"
    assert streamed_reasoning == "Inspect the prompt before replying."
    assert stream_chunks[-1]["reasoning"] == {
        "visibility": "raw_model_emitted",
        "available": True,
        "content": "Inspect the prompt before replying.",
        "summary": None,
    }


def test_end_to_end_http_attachment_aware_chat_workflow(
    app_with_fake_attachment_runtime,
    sample_attachment_sources,
) -> None:
    with TestClient(app_with_fake_attachment_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        assert scan_response.status_code == 200
        vision_model_id = next(
            manifest["model_id"]
            for manifest in scan_response.json()["manifests"]
            if manifest["display_name"] == "qwen2-vl-vision-mlx"
        )

        chat_response = client.post(
            "/v1/chat/completions",
            json={
                "messages": [
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": "Review these local inputs."},
                            {"type": "input_image", "path": str(sample_attachment_sources["image_one"])},
                            {"type": "input_file", "path": str(sample_attachment_sources["docx"])},
                            {"type": "input_audio", "path": str(sample_attachment_sources["audio"])},
                        ],
                    },
                ],
            },
        )

        response_api = client.post(
            "/v1/responses",
            json={
                "input": [
                    {
                        "role": "user",
                        "content": [
                            {"type": "input_text", "text": "Summarize the attached PDF."},
                            {"type": "input_file", "path": str(sample_attachment_sources["pdf"])},
                        ],
                    },
                ],
            },
        )

    assert chat_response.status_code == 200
    assert chat_response.json()["model"] == vision_model_id
    assert "[Attached image: receipt-front.png]" in chat_response.json()["choices"][0]["message"]["content"]
    assert "[Attached document: sample.docx]" in chat_response.json()["choices"][0]["message"]["content"]
    assert "Transcribed voice-note.wav" in chat_response.json()["choices"][0]["message"]["content"]
    assert "[images: receipt-front.png]" in chat_response.json()["choices"][0]["message"]["content"]

    assert response_api.status_code == 200
    assert "[Attached document: sample.pdf]" in response_api.json()["output_text"]


def test_end_to_end_http_multipart_multimodal_workflow(
    app_with_fake_attachment_runtime,
    sample_attachment_sources,
    sample_audio_bytes: bytes,
) -> None:
    with TestClient(app_with_fake_attachment_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        manifests = scan_response.json()["manifests"]
        vision_model_id = next(
            manifest["model_id"]
            for manifest in manifests
            if manifest["display_name"] == "qwen2-vl-vision-mlx"
        )

        chat_response = client.post(
            "/v1/chat/completions",
            data={
                "payload_json": json.dumps(
                    {
                        "messages": [
                            {
                                "role": "user",
                                "content": [
                                    {"type": "text", "text": "Inspect these uploaded assets."},
                                    {"type": "input_image", "upload_name": "image_upload"},
                                    {"type": "input_file", "upload_name": "document_upload"},
                                    {"type": "input_audio", "upload_name": "audio_upload"},
                                ],
                            },
                        ],
                    },
                ),
            },
            files=[
                ("image_upload", ("receipt-front.png", sample_attachment_sources["image_one"].read_bytes(), "image/png")),
                ("document_upload", ("sample.docx", sample_attachment_sources["docx"].read_bytes(), "application/vnd.openxmlformats-officedocument.wordprocessingml.document")),
                ("audio_upload", ("voice-note.wav", sample_audio_bytes, "audio/wav")),
            ],
        )

    assert chat_response.status_code == 200
    assert chat_response.json()["model"] == vision_model_id
    assert "[Attached image: 0-receipt-front.png]" in chat_response.json()["choices"][0]["message"]["content"]
    assert "[Attached document: 1-sample.docx]" in chat_response.json()["choices"][0]["message"]["content"]
    assert "Transcribed 2-voice-note.wav" in chat_response.json()["choices"][0]["message"]["content"]


def test_end_to_end_cli_scan_chat_warm_and_unload(
    temp_settings,
    services_with_fake_runtime,
    capsys,
) -> None:
    scan_code = main(["scan", "--json"], settings=temp_settings, services=services_with_fake_runtime)
    scan_output = json.loads(capsys.readouterr().out)
    assert scan_code == 0

    gguf_model_id = next(
        manifest["model_id"]
        for manifest in scan_output["manifests"]
        if manifest["format_type"] == "gguf"
    )

    warm_code = main(["warm", gguf_model_id, "--json"], settings=temp_settings, services=services_with_fake_runtime)
    warm_output = json.loads(capsys.readouterr().out)
    assert warm_code == 0
    assert warm_output["status"] == "warmed"

    capabilities_code = main(
        ["capabilities", gguf_model_id, "--json"],
        settings=temp_settings,
        services=services_with_fake_runtime,
    )
    capabilities_output = json.loads(capsys.readouterr().out)
    assert capabilities_code == 0
    assert capabilities_output["model_id"] == gguf_model_id
    assert any(item["capability"] == "chat" and item["supported"] is True for item in capabilities_output["capabilities"])

    chat_code = main(
        ["chat", "Draft a reply", "--model", gguf_model_id, "--json"],
        settings=temp_settings,
        services=services_with_fake_runtime,
    )
    chat_output = json.loads(capsys.readouterr().out)
    assert chat_code == 0
    assert chat_output["output_text"] == "Echo: Draft a reply"

    doctor_code = main(["doctor", "--json"], settings=temp_settings, services=services_with_fake_runtime)
    doctor_output = json.loads(capsys.readouterr().out)
    assert doctor_code == 0
    assert doctor_output["runtime_stats"]["request_metrics"]["total_requests"] >= 1
    assert doctor_output["runtime_stats"]["request_metrics"]["models"][0]["model_id"] == gguf_model_id
    assert doctor_output["runtime_stats"]["request_metrics"]["capabilities"][0]["capability"] == "chat"
    assert ("Linux", "x86_64") in {
        (item["system"], item["machine"])
        for item in doctor_output["runtime_stats"]["target_platforms"]
    }

    unload_code = main(["unload", gguf_model_id, "--json"], settings=temp_settings, services=services_with_fake_runtime)
    unload_output = json.loads(capsys.readouterr().out)
    assert unload_code == 0
    assert unload_output["status"] == "unloaded"


def test_end_to_end_cli_chat_supports_system_prompt_file(
    temp_settings,
    services_with_fake_runtime,
    tmp_path: Path,
    capsys,
) -> None:
    scan_code = main(["scan", "--json"], settings=temp_settings, services=services_with_fake_runtime)
    scan_output = json.loads(capsys.readouterr().out)
    assert scan_code == 0

    gguf_model_id = next(
        manifest["model_id"]
        for manifest in scan_output["manifests"]
        if manifest["format_type"] == "gguf"
    )
    system_prompt_path = tmp_path / "pretext.txt"
    system_prompt_path.write_text("You are a concise local assistant.", encoding="utf-8")

    chat_code = main(
        [
            "chat",
            "Respond briefly",
            "--model",
            gguf_model_id,
            "--system-prompt-file",
            str(system_prompt_path),
            "--json",
        ],
        settings=temp_settings,
        services=services_with_fake_runtime,
    )
    chat_output = json.loads(capsys.readouterr().out)
    assert chat_code == 0
    assert chat_output["output_text"] == "Echo: Respond briefly"
    assert chat_output["message_count"] == 2


def test_end_to_end_cli_chat_can_save_and_resume_session(
    session_enabled_settings,
    services_with_fake_runtime_session_enabled,
    capsys,
) -> None:
    scan_code = main(
        ["scan", "--json"],
        settings=session_enabled_settings,
        services=services_with_fake_runtime_session_enabled,
    )
    scan_output = json.loads(capsys.readouterr().out)
    assert scan_code == 0

    gguf_model_id = next(
        manifest["model_id"]
        for manifest in scan_output["manifests"]
        if manifest["format_type"] == "gguf"
    )

    first_chat_code = main(
        ["chat", "Save this note", "--model", gguf_model_id, "--save-session", "--session-title", "Roadmap", "--json"],
        settings=session_enabled_settings,
        services=services_with_fake_runtime_session_enabled,
    )
    first_chat_payload = json.loads(capsys.readouterr().out)
    assert first_chat_code == 0
    assert first_chat_payload["session_id"]

    second_chat_code = main(
        [
            "chat",
            "Continue the session",
            "--model",
            gguf_model_id,
            "--session-id",
            first_chat_payload["session_id"],
            "--json",
        ],
        settings=session_enabled_settings,
        services=services_with_fake_runtime_session_enabled,
    )
    second_chat_payload = json.loads(capsys.readouterr().out)
    assert second_chat_code == 0
    assert second_chat_payload["session_id"] == first_chat_payload["session_id"]
    assert second_chat_payload["message_count"] == 3


def test_end_to_end_cli_stream_chat_can_save_session_without_crashing(
    session_enabled_settings,
    services_with_fake_runtime_session_enabled,
    capsys,
) -> None:
    scan_code = main(
        ["scan", "--json"],
        settings=session_enabled_settings,
        services=services_with_fake_runtime_session_enabled,
    )
    scan_output = json.loads(capsys.readouterr().out)
    assert scan_code == 0

    gguf_model_id = next(
        manifest["model_id"]
        for manifest in scan_output["manifests"]
        if manifest["format_type"] == "gguf"
    )

    stream_code = main(
        [
            "chat",
            "Stream and save this note",
            "--model",
            gguf_model_id,
            "--save-session",
            "--session-title",
            "Streaming Notes",
            "--stream",
        ],
        settings=session_enabled_settings,
        services=services_with_fake_runtime_session_enabled,
    )
    stream_output = capsys.readouterr().out
    assert stream_code == 0
    assert "Echo: Stream and save this note" in stream_output
    assert "session: " in stream_output

    list_code = main(
        ["list-sessions", "--json"],
        settings=session_enabled_settings,
        services=services_with_fake_runtime_session_enabled,
    )
    list_output = json.loads(capsys.readouterr().out)
    assert list_code == 0
    assert list_output["count"] == 1
    assert list_output["items"][0]["title"] == "Streaming Notes"


def test_end_to_end_cli_session_management_commands(
    session_enabled_settings,
    services_with_fake_runtime_session_enabled,
    tmp_path: Path,
    capsys,
) -> None:
    scan_code = main(
        ["scan", "--json"],
        settings=session_enabled_settings,
        services=services_with_fake_runtime_session_enabled,
    )
    scan_output = json.loads(capsys.readouterr().out)
    assert scan_code == 0

    gguf_model_id = next(
        manifest["model_id"]
        for manifest in scan_output["manifests"]
        if manifest["format_type"] == "gguf"
    )

    chat_code = main(
        ["chat", "Persist this session", "--model", gguf_model_id, "--save-session", "--session-title", "Ops Notes", "--json"],
        settings=session_enabled_settings,
        services=services_with_fake_runtime_session_enabled,
    )
    chat_output = json.loads(capsys.readouterr().out)
    assert chat_code == 0
    session_id = chat_output["session_id"]

    list_code = main(
        ["list-sessions", "--json"],
        settings=session_enabled_settings,
        services=services_with_fake_runtime_session_enabled,
    )
    list_output = json.loads(capsys.readouterr().out)
    assert list_code == 0
    assert list_output["count"] == 1
    assert list_output["items"][0]["session_id"] == session_id

    show_code = main(
        ["show-session", session_id, "--json"],
        settings=session_enabled_settings,
        services=services_with_fake_runtime_session_enabled,
    )
    show_output = json.loads(capsys.readouterr().out)
    assert show_code == 0
    assert show_output["title"] == "Ops Notes"
    assert show_output["turn_count"] == 1
    assert show_output["turns"][0]["response_message"]["content"] == "Echo: Persist this session"

    export_path = tmp_path / "session.json"
    export_code = main(
        ["export-session", session_id, "--output", str(export_path), "--json"],
        settings=session_enabled_settings,
        services=services_with_fake_runtime_session_enabled,
    )
    export_output = json.loads(capsys.readouterr().out)
    assert export_code == 0
    assert export_output["output_path"] == str(export_path)
    assert export_path.exists()

    import_code = main(
        ["import-session", "--input", str(export_path), "--title", "Imported Ops Notes", "--json"],
        settings=session_enabled_settings,
        services=services_with_fake_runtime_session_enabled,
    )
    import_output = json.loads(capsys.readouterr().out)
    assert import_code == 0
    assert import_output["session_id"] != session_id
    assert import_output["title"] == "Imported Ops Notes"
    assert import_output["turn_count"] == 1

    delete_code = main(
        ["delete-session", session_id, "--json"],
        settings=session_enabled_settings,
        services=services_with_fake_runtime_session_enabled,
    )
    delete_output = json.loads(capsys.readouterr().out)
    assert delete_code == 0
    assert delete_output["status"] == "deleted"

    final_list_code = main(
        ["list-sessions", "--json"],
        settings=session_enabled_settings,
        services=services_with_fake_runtime_session_enabled,
    )
    final_list_output = json.loads(capsys.readouterr().out)
    assert final_list_code == 0
    assert final_list_output["count"] == 1
    assert final_list_output["items"][0]["session_id"] == import_output["session_id"]


def test_end_to_end_cli_skill_and_tool_catalog_commands(
    temp_settings,
    contract_transform_payload: dict[str, object],
    tmp_path: Path,
    capsys,
) -> None:
    list_skills_code = main(["list-skills", "--json"], settings=temp_settings)
    list_skills_payload = json.loads(capsys.readouterr().out)
    assert list_skills_code == 0
    assert any(item["name"] == "document_comparison" for item in list_skills_payload["items"])

    show_skill_code = main(["show-skill", "document_comparison", "--json"], settings=temp_settings)
    show_skill_payload = json.loads(capsys.readouterr().out)
    assert show_skill_code == 0
    assert show_skill_payload["tool_name"] == "documents.transform"

    list_tools_code = main(["list-tools", "--json"], settings=temp_settings)
    list_tools_payload = json.loads(capsys.readouterr().out)
    assert list_tools_code == 0
    assert any(item["name"] == "documents.transform" for item in list_tools_payload["items"])

    show_tool_code = main(["show-tool", "documents.transform", "--json"], settings=temp_settings)
    show_tool_payload = json.loads(capsys.readouterr().out)
    assert show_tool_code == 0
    assert show_tool_payload["required_authorization"] == "document_transform"

    request_path = tmp_path / "tool-transform.json"
    request_path.write_text(
        json.dumps({"tool": "documents.transform", "input": contract_transform_payload}, indent=2),
        encoding="utf-8",
    )
    output_path = tmp_path / "tool-contract.docx"

    run_tool_code = main(
        [
            "run-tool",
            "--input",
            str(request_path),
            "--output",
            str(output_path),
            "--json",
        ],
        settings=temp_settings,
    )
    run_tool_payload = json.loads(capsys.readouterr().out)
    assert run_tool_code == 0
    assert run_tool_payload["request_id"]
    assert run_tool_payload["tool"] == "documents.transform"
    assert run_tool_payload["result"]["output_path"] == str(output_path)
    assert output_path.read_bytes().startswith(b"PK")


def test_end_to_end_cli_chat_supports_session_context_policies(
    session_enabled_settings,
    services_with_fake_runtime_session_enabled,
    capsys,
) -> None:
    scan_code = main(
        ["scan", "--json"],
        settings=session_enabled_settings,
        services=services_with_fake_runtime_session_enabled,
    )
    scan_output = json.loads(capsys.readouterr().out)
    assert scan_code == 0

    gguf_model_id = next(
        manifest["model_id"]
        for manifest in scan_output["manifests"]
        if manifest["format_type"] == "gguf"
    )

    first_chat_code = main(
        [
            "chat",
            "First compacted note",
            "--model",
            gguf_model_id,
            "--save-session",
            "--session-title",
            "Compacted Roadmap",
            "--session-context-policy",
            "last_turn",
            "--json",
        ],
        settings=session_enabled_settings,
        services=services_with_fake_runtime_session_enabled,
    )
    first_chat_payload = json.loads(capsys.readouterr().out)
    assert first_chat_code == 0

    second_chat_code = main(
        [
            "chat",
            "Second compacted note",
            "--model",
            gguf_model_id,
            "--session-id",
            first_chat_payload["session_id"],
            "--json",
        ],
        settings=session_enabled_settings,
        services=services_with_fake_runtime_session_enabled,
    )
    second_chat_payload = json.loads(capsys.readouterr().out)
    assert second_chat_code == 0
    assert second_chat_payload["message_count"] == 3

    third_chat_code = main(
        [
            "chat",
            "Third compacted note",
            "--model",
            gguf_model_id,
            "--session-id",
            first_chat_payload["session_id"],
            "--json",
        ],
        settings=session_enabled_settings,
        services=services_with_fake_runtime_session_enabled,
    )
    third_chat_payload = json.loads(capsys.readouterr().out)
    assert third_chat_code == 0
    assert third_chat_payload["message_count"] == 3

    show_code = main(
        ["show-session", first_chat_payload["session_id"], "--json"],
        settings=session_enabled_settings,
        services=services_with_fake_runtime_session_enabled,
    )
    show_payload = json.loads(capsys.readouterr().out)
    assert show_code == 0
    assert show_payload["context_policy"] == "last_turn"
    assert show_payload["turn_count"] == 3


def test_end_to_end_cli_chat_supports_prompt_override_files(
    temp_settings,
    services_with_fake_runtime,
    sample_prompt_assets,
    capsys,
) -> None:
    scan_code = main(["scan", "--json"], settings=temp_settings, services=services_with_fake_runtime)
    scan_output = json.loads(capsys.readouterr().out)
    assert scan_code == 0

    gguf_model_id = next(
        manifest["model_id"]
        for manifest in scan_output["manifests"]
        if manifest["format_type"] == "gguf"
    )

    chat_code = main(
        [
            "chat",
            "Return the milestone summary",
            "--model",
            gguf_model_id,
            "--developer-prompt",
            "Keep it terse.",
            "--pretext-file",
            str(sample_prompt_assets["pretext"]),
            "--skills-file",
            str(sample_prompt_assets["skill"]),
            "--response-format-file",
            str(sample_prompt_assets["response_format"]),
            "--tools-file",
            str(sample_prompt_assets["tools"]),
            "--mcp-tools-file",
            str(sample_prompt_assets["mcp_tools"]),
            "--json",
        ],
        settings=temp_settings,
        services=services_with_fake_runtime,
    )
    chat_output = json.loads(capsys.readouterr().out)
    assert chat_code == 0
    assert chat_output["output_text"] == '{"summary":"ok","status":"ok"}'
    assert chat_output["prompt_trace"]["selected_template"] == "tool_structured_output"
    assert chat_output["prompt_trace"]["model_prompt_template"]["id"] == "llama-instruct-v1"
    assert "[INST]" in chat_output["prompt_trace"]["serialized_model_prompt"]
    assert [override["source"] for override in chat_output["prompt_trace"]["overrides"]] == [
        "pretext_file",
        "developer_prompt",
        "skills_file",
        "tools_file",
        "mcp_tools_file",
        "response_format_file",
    ]
    mcp_tool = next(tool for tool in chat_output["prompt_trace"]["tool_plan"] if tool["name"] == "search_milestones")
    assert mcp_tool["mcp_server"] == "roadmap"
    assert mcp_tool["metadata_trusted"] is False


def test_end_to_end_cli_chat_supports_local_attachments(
    temp_settings,
    services_with_fake_attachment_runtime,
    sample_attachment_sources,
    capsys,
) -> None:
    scan_code = main(["scan", "--json"], settings=temp_settings, services=services_with_fake_attachment_runtime)
    scan_output = json.loads(capsys.readouterr().out)
    assert scan_code == 0

    vision_model_id = next(
        manifest["model_id"]
        for manifest in scan_output["manifests"]
        if manifest["display_name"] == "qwen2-vl-vision-mlx"
    )

    chat_code = main(
        [
            "chat",
            "Review these local inputs.",
            "--attach-image",
            str(sample_attachment_sources["image_one"]),
            "--attach-file",
            str(sample_attachment_sources["pdf"]),
            "--attach-audio",
            str(sample_attachment_sources["audio"]),
            "--json",
        ],
        settings=temp_settings,
        services=services_with_fake_attachment_runtime,
    )
    chat_output = json.loads(capsys.readouterr().out)
    assert chat_code == 0
    assert chat_output["model"] == vision_model_id
    assert "[Attached image: receipt-front.png]" in chat_output["output_text"]
    assert "[Attached document: sample.pdf]" in chat_output["output_text"]
    assert "Transcribed voice-note.wav" in chat_output["output_text"]


def test_end_to_end_cli_generate_document_files(
    temp_settings,
    sample_document_ir,
    tmp_path: Path,
    capsys,
) -> None:
    input_path = tmp_path / "document.json"
    input_path.write_text(sample_document_ir.model_dump_json(indent=2), encoding="utf-8")

    outputs = {
        "text": tmp_path / "report.txt",
        "markdown": tmp_path / "report.md",
        "json": tmp_path / "report.json",
        "csv": tmp_path / "report.csv",
        "docx": tmp_path / "report.docx",
        "pdf": tmp_path / "report.pdf",
        "xlsx": tmp_path / "report.xlsx",
    }

    for output_format, output_path in outputs.items():
        exit_code = main(
            [
                "generate-doc",
                "--input",
                str(input_path),
                "--format",
                output_format,
                "--output",
                str(output_path),
                "--json",
            ],
            settings=temp_settings,
        )
        payload = json.loads(capsys.readouterr().out)
        assert exit_code == 0
        assert payload["output_path"] == str(output_path)
        assert output_path.exists()

    assert outputs["text"].read_text(encoding="utf-8").startswith("Quarterly Operations Summary")
    assert outputs["markdown"].read_text(encoding="utf-8").startswith("# Quarterly Operations Summary")
    assert json.loads(outputs["json"].read_text(encoding="utf-8"))["title"] == "Quarterly Operations Summary"
    assert outputs["csv"].read_text(encoding="utf-8").startswith("Category,Amount")
    assert outputs["docx"].read_bytes().startswith(b"PK")
    assert outputs["pdf"].read_bytes().startswith(b"%PDF")
    workbook = load_workbook(outputs["xlsx"])
    assert workbook["Budget"]["A4"].value == "Category"


def test_end_to_end_cli_generate_requires_authorization(
    tool_authorized_settings,
    sample_document_ir,
    tmp_path: Path,
    capsys,
) -> None:
    input_path = tmp_path / "document.json"
    input_path.write_text(sample_document_ir.model_dump_json(indent=2), encoding="utf-8")
    output_path = tmp_path / "secured-report.csv"

    denied_code = main(
        [
            "generate-doc",
            "--input",
            str(input_path),
            "--format",
            "csv",
            "--output",
            str(output_path),
            "--json",
        ],
        settings=tool_authorized_settings,
    )
    denied_payload = json.loads(capsys.readouterr().out)
    assert denied_code == 1
    assert denied_payload["error"]["code"] == "tool_authorization_error"

    allowed_code = main(
        [
            "generate-doc",
            "--input",
            str(input_path),
            "--format",
            "csv",
            "--output",
            str(output_path),
            "--authorize",
            "document_generate",
            "--json",
        ],
        settings=tool_authorized_settings,
    )
    allowed_payload = json.loads(capsys.readouterr().out)
    assert allowed_code == 0
    assert allowed_payload["output_path"] == str(output_path)
    assert output_path.exists()


def test_end_to_end_cli_transform_skill_requests(
    temp_settings,
    contract_transform_payload: dict[str, object],
    receipt_transform_payload: dict[str, object],
    branded_document_template_payload: dict[str, object],
    ocr_assisted_extraction_payload: dict[str, object],
    file_template_transform_payload: dict[str, object],
    document_compare_transform_payload: dict[str, object],
    meeting_transcript_notes_payload: dict[str, object],
    long_document_memo_payload: dict[str, object],
    speech_transcript_cleanup_payload: dict[str, object],
    tmp_path: Path,
    capsys,
) -> None:
    contract_request_path = tmp_path / "contract-transform.json"
    contract_request_path.write_text(json.dumps(contract_transform_payload, indent=2), encoding="utf-8")
    contract_output_path = tmp_path / "service-agreement.docx"

    contract_code = main(
        [
            "transform",
            "--input",
            str(contract_request_path),
            "--output",
            str(contract_output_path),
            "--json",
        ],
        settings=temp_settings,
    )
    contract_payload = json.loads(capsys.readouterr().out)
    assert contract_code == 0
    assert contract_payload["skill"] == "contract_text_replacement"
    assert contract_output_path.read_bytes().startswith(b"PK")

    receipt_request_path = tmp_path / "receipt-transform.json"
    receipt_request_path.write_text(json.dumps(receipt_transform_payload, indent=2), encoding="utf-8")
    receipt_output_path = tmp_path / "receipt.csv"

    receipt_code = main(
        [
            "transform",
            "--input",
            str(receipt_request_path),
            "--output",
            str(receipt_output_path),
            "--json",
        ],
        settings=temp_settings,
    )
    receipt_payload = json.loads(capsys.readouterr().out)
    assert receipt_code == 0
    assert receipt_payload["skill"] == "receipt_extraction"
    assert receipt_output_path.read_text(encoding="utf-8").startswith("Description,Quantity,Unit Price,Line Total")

    branding_assets_dir = tmp_path / "assets"
    branding_assets_dir.mkdir(parents=True, exist_ok=True)
    logo_source_path = Path(branded_document_template_payload["input"]["settings"]["logo_path"])
    hero_source_path = Path(branded_document_template_payload["input"]["settings"]["hero_image_path"])
    logo_target_path = branding_assets_dir / logo_source_path.name
    hero_target_path = branding_assets_dir / hero_source_path.name
    logo_target_path.write_bytes(logo_source_path.read_bytes())
    hero_target_path.write_bytes(hero_source_path.read_bytes())
    branded_request_payload = json.loads(json.dumps(branded_document_template_payload))
    branded_request_payload["input"]["settings"]["logo_path"] = f"assets/{logo_target_path.name}"
    branded_request_payload["input"]["settings"]["hero_image_path"] = f"assets/{hero_target_path.name}"
    branded_request_path = tmp_path / "branded-document-template.json"
    branded_request_path.write_text(json.dumps(branded_request_payload, indent=2), encoding="utf-8")
    branded_output_path = tmp_path / "product-launch-brief.json"

    branded_code = main(
        [
            "transform",
            "--input",
            str(branded_request_path),
            "--output",
            str(branded_output_path),
            "--json",
        ],
        settings=temp_settings,
    )
    branded_payload = json.loads(capsys.readouterr().out)
    assert branded_code == 0
    assert branded_payload["skill"] == "branded_document_template"
    branded_document = json.loads(branded_output_path.read_text(encoding="utf-8"))
    assert branded_document["header"]["left"] == "LewLM"
    assert branded_document["footer"]["right"] == "Leadership Team"
    image_blocks = [
        block
        for section in branded_document["sections"]
        for block in section["blocks"]
        if block["type"] == "image"
    ]
    assert len(image_blocks) == 2
    assert any(block["role"] == "logo" for block in image_blocks)

    ocr_request_path = tmp_path / "ocr-assisted-extraction.json"
    ocr_request_path.write_text(json.dumps(ocr_assisted_extraction_payload, indent=2), encoding="utf-8")
    ocr_output_path = tmp_path / "scanned-invoice-extraction.md"

    ocr_code = main(
        [
            "transform",
            "--input",
            str(ocr_request_path),
            "--output",
            str(ocr_output_path),
            "--json",
        ],
        settings=temp_settings,
    )
    ocr_payload = json.loads(capsys.readouterr().out)
    assert ocr_code == 0
    assert ocr_payload["skill"] == "ocr_assisted_extraction"
    ocr_markdown = ocr_output_path.read_text(encoding="utf-8")
    assert ocr_markdown.startswith("# Scanned Invoice Extraction")
    assert "## Review Notes" in ocr_markdown
    assert "Total Due | USD 1,240.00 | extracted" in ocr_markdown

    template_source_path = Path(file_template_transform_payload["template_path"])
    local_template_path = tmp_path / template_source_path.name
    local_template_path.write_text(template_source_path.read_text(encoding="utf-8"), encoding="utf-8")
    template_request_payload = dict(file_template_transform_payload)
    template_request_payload["template_path"] = local_template_path.name
    template_request_path = tmp_path / "template-transform.json"
    template_request_path.write_text(json.dumps(template_request_payload, indent=2), encoding="utf-8")
    template_output_path = tmp_path / "engagement-summary.xlsx"

    template_code = main(
        [
            "transform",
            "--input",
            str(template_request_path),
            "--output",
            str(template_output_path),
            "--json",
        ],
        settings=temp_settings,
    )
    template_payload = json.loads(capsys.readouterr().out)
    assert template_code == 0
    assert template_payload["skill"] == "file_template"
    workbook = load_workbook(template_output_path)
    assert workbook["Overview"]["B9"].value == "On Track"

    compare_request_path = tmp_path / "compare-transform.json"
    compare_request_path.write_text(json.dumps(document_compare_transform_payload, indent=2), encoding="utf-8")
    compare_output_path = tmp_path / "agreement-comparison.xlsx"

    compare_code = main(
        [
            "transform",
            "--input",
            str(compare_request_path),
            "--output",
            str(compare_output_path),
            "--json",
        ],
        settings=temp_settings,
    )
    compare_payload = json.loads(capsys.readouterr().out)
    assert compare_code == 0
    assert compare_payload["skill"] == "document_comparison"
    comparison_workbook = load_workbook(compare_output_path)
    assert comparison_workbook["Overview"]["B8"].value == "Baseline Agreement"
    assert comparison_workbook["Shared Segments"]["A3"].value == "- Shared scope line."

    meeting_request_path = tmp_path / "meeting-notes.json"
    meeting_request_path.write_text(json.dumps(meeting_transcript_notes_payload, indent=2), encoding="utf-8")
    meeting_output_path = tmp_path / "meeting-notes.md"

    meeting_code = main(
        [
            "transform",
            "--input",
            str(meeting_request_path),
            "--output",
            str(meeting_output_path),
            "--json",
        ],
        settings=temp_settings,
    )
    meeting_payload = json.loads(capsys.readouterr().out)
    assert meeting_code == 0
    assert meeting_payload["skill"] == "meeting_transcript_notes"
    meeting_markdown = meeting_output_path.read_text(encoding="utf-8")
    assert meeting_markdown.startswith("# Project Kickoff Notes")
    assert "## Action Items" in meeting_markdown
    assert "Jon | Prepare rollout checklist | 2026-04-22" in meeting_markdown

    memo_request_path = tmp_path / "long-document-memo.json"
    memo_request_path.write_text(json.dumps(long_document_memo_payload, indent=2), encoding="utf-8")
    memo_output_path = tmp_path / "platform-readiness-memo.md"

    memo_code = main(
        [
            "transform",
            "--input",
            str(memo_request_path),
            "--output",
            str(memo_output_path),
            "--json",
        ],
        settings=temp_settings,
    )
    memo_payload = json.loads(capsys.readouterr().out)
    assert memo_code == 0
    assert memo_payload["skill"] == "long_document_memo"
    memo_markdown = memo_output_path.read_text(encoding="utf-8")
    assert memo_markdown.startswith("# Platform Readiness Memo")
    assert "## Memo Summary" in memo_markdown
    assert "## Source Outline" in memo_markdown

    cleanup_request_path = tmp_path / "speech-transcript-cleanup.json"
    cleanup_request_path.write_text(json.dumps(speech_transcript_cleanup_payload, indent=2), encoding="utf-8")
    cleanup_output_path = tmp_path / "customer-call-cleanup.md"

    cleanup_code = main(
        [
            "transform",
            "--input",
            str(cleanup_request_path),
            "--output",
            str(cleanup_output_path),
            "--json",
        ],
        settings=temp_settings,
    )
    cleanup_payload = json.loads(capsys.readouterr().out)
    assert cleanup_code == 0
    assert cleanup_payload["skill"] == "speech_transcript_cleanup"
    cleanup_markdown = cleanup_output_path.read_text(encoding="utf-8")
    assert cleanup_markdown.startswith("# Customer Call Cleanup")
    assert "## Cleaned Transcript" in cleanup_markdown
    assert "Customer | I have two questions about rollout timing." in cleanup_markdown


def test_end_to_end_cli_convert_cache_and_benchmark(
    temp_settings,
    services_with_fake_runtime_and_conversion,
    capsys,
) -> None:
    scan_code = main(["scan", "--json"], settings=temp_settings, services=services_with_fake_runtime_and_conversion)
    scan_payload = json.loads(capsys.readouterr().out)
    assert scan_code == 0

    gguf_model_id = next(
        manifest["model_id"]
        for manifest in scan_payload["manifests"]
        if manifest["format_type"] == "gguf"
    )
    hf_model_id = next(
        manifest["model_id"]
        for manifest in scan_payload["manifests"]
        if manifest["format_type"] == "huggingface"
    )

    convert_code = main(
        ["convert", hf_model_id, "--idempotency-key", "cli-convert-1", "--json"],
        settings=temp_settings,
        services=services_with_fake_runtime_and_conversion,
    )
    convert_payload = json.loads(capsys.readouterr().out)
    assert convert_code == 0
    assert convert_payload["status"] == "completed"
    assert convert_payload["idempotency_key"] == "cli-convert-1"
    assert Path(convert_payload["payload"]["result_path"]).exists()

    repeat_convert_code = main(
        ["convert", hf_model_id, "--idempotency-key", "cli-convert-1", "--json"],
        settings=temp_settings,
        services=services_with_fake_runtime_and_conversion,
    )
    repeat_convert_payload = json.loads(capsys.readouterr().out)
    assert repeat_convert_code == 0
    assert repeat_convert_payload["job_id"] == convert_payload["job_id"]
    assert repeat_convert_payload["idempotency_key"] == "cli-convert-1"
    assert repeat_convert_payload["idempotent_replay"] is True

    cache_repeat_convert_code = main(
        ["convert", hf_model_id, "--json"],
        settings=temp_settings,
        services=services_with_fake_runtime_and_conversion,
    )
    cache_repeat_convert_payload = json.loads(capsys.readouterr().out)
    assert cache_repeat_convert_code == 0
    assert cache_repeat_convert_payload["payload"]["cache_hit"] is True

    cache_code = main(["cache", "--json"], settings=temp_settings, services=services_with_fake_runtime_and_conversion)
    cache_payload = json.loads(capsys.readouterr().out)
    assert cache_code == 0
    assert cache_payload["artifact_count"] == 1

    benchmark_code = main(
        ["benchmark", "--model", gguf_model_id, "--json"],
        settings=temp_settings,
        services=services_with_fake_runtime_and_conversion,
    )
    benchmark_payload = json.loads(capsys.readouterr().out)
    assert benchmark_code == 0
    assert benchmark_payload["runtime"] == "fake_llamacpp"
    assert benchmark_payload["output_text"] == "Echo: Benchmark ping"
    assert benchmark_payload["benchmark_id"]
    assert benchmark_payload["created_at"]

    doctor_code = main(["doctor", "--json"], settings=temp_settings, services=services_with_fake_runtime_and_conversion)
    doctor_payload = json.loads(capsys.readouterr().out)
    assert doctor_code == 0
    assert doctor_payload["runtime_stats"]["benchmark_summary"]["total_runs"] == 1
    assert doctor_payload["runtime_stats"]["benchmark_summary"]["recent_runs"][0]["model_id"] == gguf_model_id


def test_end_to_end_cli_benchmark_all_models(
    temp_settings,
    services_with_fake_runtime_and_conversion,
    capsys,
) -> None:
    second_model_path = temp_settings.models_dir[0] / "mistral-7b-instruct-q4_k_m.gguf"
    second_model_path.write_bytes(b"gguf-model-2")

    scan_code = main(["scan", "--json"], settings=temp_settings, services=services_with_fake_runtime_and_conversion)
    scan_payload = json.loads(capsys.readouterr().out)
    assert scan_code == 0
    assert sum(1 for manifest in scan_payload["manifests"] if manifest["format_type"] == "gguf") == 2

    benchmark_code = main(
        ["benchmark", "--all", "--json"],
        settings=temp_settings,
        services=services_with_fake_runtime_and_conversion,
    )
    benchmark_payload = json.loads(capsys.readouterr().out)
    assert benchmark_code == 0
    assert benchmark_payload["benchmark_count"] == 2
    assert benchmark_payload["model_count"] == 2
    assert benchmark_payload["repeat_count"] == 1
    assert len(benchmark_payload["results"]) == 2
    assert len(benchmark_payload["models"]) == 2

    doctor_code = main(["doctor", "--json"], settings=temp_settings, services=services_with_fake_runtime_and_conversion)
    doctor_payload = json.loads(capsys.readouterr().out)
    assert doctor_code == 0
    assert doctor_payload["runtime_stats"]["benchmark_summary"]["total_runs"] == 2
    assert len(doctor_payload["runtime_stats"]["benchmark_summary"]["models"]) == 2


def test_end_to_end_cli_benchmark_all_embedding_models(
    temp_settings,
    services_with_fake_multimodal_runtime,
    capsys,
) -> None:
    second_embedding_dir = temp_settings.models_dir[0] / "gte-small-embed-mlx"
    second_embedding_dir.mkdir(parents=True)
    (second_embedding_dir / "config.json").write_text(
        json.dumps({"model_type": "gte", "max_position_embeddings": 8192}),
        encoding="utf-8",
    )
    (second_embedding_dir / "weights.safetensors").write_bytes(b"embed-weights-2")
    (second_embedding_dir / "tokenizer.json").write_text("{}", encoding="utf-8")

    scan_code = main(["scan", "--json"], settings=temp_settings, services=services_with_fake_multimodal_runtime)
    scan_payload = json.loads(capsys.readouterr().out)
    assert scan_code == 0
    expected_model_count = sum(
        1
        for manifest in scan_payload["manifests"]
        if any(
            capability["capability"] == "embeddings" and capability["supported"] is True
            for capability in services_with_fake_multimodal_runtime.model_router.model_capability_report(
                manifest["model_id"],
            ).model_dump(mode="json")["capabilities"]
        )
    )
    assert expected_model_count >= 2

    benchmark_code = main(
        ["benchmark", "--all", "--capability", "embeddings", "--json"],
        settings=temp_settings,
        services=services_with_fake_multimodal_runtime,
    )
    benchmark_payload = json.loads(capsys.readouterr().out)
    assert benchmark_code == 0
    assert benchmark_payload["capability"] == "embeddings"
    assert benchmark_payload["benchmark_count"] == expected_model_count
    assert benchmark_payload["model_count"] == expected_model_count
    assert benchmark_payload["repeat_count"] == 1
    assert len(benchmark_payload["results"]) == expected_model_count
    assert all(item["capability"] == "embeddings" for item in benchmark_payload["results"])

    doctor_code = main(["doctor", "--json"], settings=temp_settings, services=services_with_fake_multimodal_runtime)
    doctor_payload = json.loads(capsys.readouterr().out)
    assert doctor_code == 0
    assert doctor_payload["runtime_stats"]["benchmark_summary"]["total_runs"] == expected_model_count
    assert doctor_payload["runtime_stats"]["benchmark_summary"]["capability_counts"]["embeddings"] == expected_model_count


def test_end_to_end_cli_benchmark_all_models_repeated(
    temp_settings,
    services_with_fake_runtime_and_conversion,
    capsys,
) -> None:
    second_model_path = temp_settings.models_dir[0] / "mistral-7b-instruct-q4_k_m.gguf"
    second_model_path.write_bytes(b"gguf-model-2")

    scan_code = main(["scan", "--json"], settings=temp_settings, services=services_with_fake_runtime_and_conversion)
    scan_payload = json.loads(capsys.readouterr().out)
    assert scan_code == 0
    assert sum(1 for manifest in scan_payload["manifests"] if manifest["format_type"] == "gguf") == 2

    benchmark_code = main(
        ["benchmark", "--all", "--repeat", "2", "--json"],
        settings=temp_settings,
        services=services_with_fake_runtime_and_conversion,
    )
    benchmark_payload = json.loads(capsys.readouterr().out)
    assert benchmark_code == 0
    assert benchmark_payload["benchmark_count"] == 4
    assert benchmark_payload["model_count"] == 2
    assert benchmark_payload["repeat_count"] == 2
    assert len(benchmark_payload["results"]) == 4
    assert len(benchmark_payload["models"]) == 2
    assert all(item["run_count"] == 2 for item in benchmark_payload["models"])

    doctor_code = main(["doctor", "--json"], settings=temp_settings, services=services_with_fake_runtime_and_conversion)
    doctor_payload = json.loads(capsys.readouterr().out)
    assert doctor_code == 0
    assert doctor_payload["runtime_stats"]["benchmark_summary"]["total_runs"] == 4
    assert doctor_payload["runtime_stats"]["benchmark_summary"]["capability_counts"]["chat"] == 4

from __future__ import annotations

import base64
import json
from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from lewlm.api.app import create_app
from lewlm.core.bootstrap import bootstrap_services
from lewlm.core.contracts import (
    CapabilityName,
    GenerateRequest,
    GenerateResponse,
    ModelFormat,
    ModelManifest,
    ModelModality,
    RuntimeAffinity,
)
from lewlm.runtime.base import ManagedTextRuntime

from conftest import FakeLlamaCppRuntime


def _api_upload_workspaces(temp_settings) -> list[Path]:
    return sorted(path for path in temp_settings.temp_dir.glob("api-upload-*") if path.is_dir())


def _write_external_validation_manifest(
    path: Path,
    *,
    capability_report: dict[str, object],
    system: str,
    machine: str,
) -> None:
    external_model = dict(capability_report)
    external_model["target_platforms"] = [
        {
            "system": system,
            "machine": machine,
            "supported": True,
            "readiness_state": "verified",
            "verification_method": "host_probe",
            "runtime_affinities": ["llamacpp"],
            "reason": f"Validated on {system} {machine}.",
            "fallback_available": False,
            "fallback_reason": None,
            "install_hints": [],
            "validation_manifest_count": 0,
            "verified_hosts": [],
            "notes": [],
        },
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(
            {
                "format": "lewlm-release-manifest-v1",
                "generated_at": "2026-04-15T00:00:00+00:00",
                "git_commit": "abc1234def5678",
                "platform": {"system": system, "machine": machine, "release": "validated-host"},
                "registered_models": [external_model],
            },
            indent=2,
        ),
        encoding="utf-8",
    )


def test_health_endpoint_reports_service_status(temp_settings) -> None:
    app = create_app(temp_settings)

    with TestClient(app) as client:
        response = client.get("/v1/health")

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "ok"
    assert payload["install_profiles"]["active_profile_ids"][0] == "core_only"
    assert any(
        item["profile"] == "external_accelerator_bridge_backend"
        for item in payload["install_profiles"]["profiles"]
    )
    assert payload["readiness"]["status"] == "blocked"
    assert payload["readiness"]["ready_capability_count"] == 0
    assert payload["readiness"]["capability_count"] >= 7
    assert payload["readiness"]["capabilities"][0]["readiness_state"] == "no_models"
    assert any(item["capability"] == "vision" for item in payload["readiness"]["capabilities"])
    assert "total_memory_mb" in payload["readiness"]["host_platform"]
    assert "total_memory_source" in payload["readiness"]["host_platform"]
    assert "total_memory_reason" in payload["readiness"]["host_platform"]
    assert payload["configuration"]["runtime_packs"]
    assert payload["configuration"]["feature_packs"]
    external_pack = next(
        item
        for item in payload["configuration"]["runtime_packs"]
        if item["name"] == "external_accelerator"
    )
    assert "loopback-only OpenAI-compatible local servers" in external_pack["description"]
    assert payload["configuration"]["privacy_mode"] is True
    assert payload["configuration"]["audit_log_enabled"] is False
    assert payload["configuration"]["persistence_encryption_enabled"] is False
    assert payload["configuration"]["tool_authorization_required"] is False
    assert payload["configuration"]["parser_sandbox_enabled"] is True
    assert payload["configuration"]["tool_sandbox_enabled"] is True
    assert payload["configuration"]["conversion_sandbox_enabled"] is True


def test_documents_feature_pack_disabled_hides_optional_surfaces(temp_settings) -> None:
    disabled_settings = temp_settings.with_updates(disabled_feature_packs=("documents",))
    app = create_app(disabled_settings)

    with TestClient(app) as client:
        health_response = client.get("/v1/health")
        skills_response = client.get("/v1/skills")
        tools_response = client.get("/v1/tools")
        generate_response = client.post(
            "/v1/documents/generate",
            json={
                "output_format": "markdown",
                "document": {
                    "title": "Disabled documents",
                    "sections": [],
                },
            },
        )

    assert health_response.status_code == 200
    feature_pack = next(
        item
        for item in health_response.json()["configuration"]["feature_packs"]
        if item["name"] == "documents"
    )
    assert feature_pack["status"] == "disabled"
    assert skills_response.status_code == 200
    assert skills_response.json() == {"count": 0, "items": []}
    assert tools_response.status_code == 200
    assert tools_response.json() == {"count": 0, "items": []}
    assert generate_response.status_code == 503
    assert generate_response.json()["error"]["code"] == "pack_unavailable"


def test_app_lifespan_closes_owned_services(temp_settings, monkeypatch) -> None:
    app = create_app(temp_settings)
    shutdown_calls: list[dict[str, object]] = []

    with TestClient(app) as client:
        services = client.app.state.services
        original_shutdown = services.conversion_service.executor.shutdown

        def wrapped_shutdown(*args, **kwargs):
            shutdown_calls.append({"args": args, "kwargs": kwargs})
            return original_shutdown(*args, **kwargs)

        monkeypatch.setattr(services.conversion_service.executor, "shutdown", wrapped_shutdown)
        response = client.get("/v1/health")

    assert response.status_code == 200
    assert shutdown_calls == [{"args": (), "kwargs": {"wait": True, "cancel_futures": True}}]


def test_app_lifespan_preserves_external_services(
    temp_settings,
    services_with_fake_runtime_and_conversion,
    monkeypatch,
) -> None:
    app = create_app(temp_settings, services=services_with_fake_runtime_and_conversion)
    shutdown_calls: list[dict[str, object]] = []
    original_shutdown = services_with_fake_runtime_and_conversion.conversion_service.executor.shutdown

    def wrapped_shutdown(*args, **kwargs):
        shutdown_calls.append({"args": args, "kwargs": kwargs})
        return original_shutdown(*args, **kwargs)

    monkeypatch.setattr(services_with_fake_runtime_and_conversion.conversion_service.executor, "shutdown", wrapped_shutdown)

    with TestClient(app) as client:
        response = client.get("/v1/health")

    assert response.status_code == 200
    assert shutdown_calls == []

    services_with_fake_runtime_and_conversion.close()


def test_model_scan_and_list_endpoints(temp_settings, sample_models_root: Path) -> None:
    app = create_app(temp_settings)

    with TestClient(app) as client:
        scan_response = client.post("/v1/models/scan", json={})
        list_response = client.get("/v1/models")

    assert scan_response.status_code == 200
    assert scan_response.json()["discovered_count"] == 3
    assert list_response.status_code == 200
    assert list_response.json()["count"] == 3


def test_model_capabilities_endpoint_reports_runtime_candidates_and_validation_manifests(
    temp_settings,
    sample_models_root: Path,
) -> None:
    baseline_services = bootstrap_services(
        temp_settings,
        runtime_overrides={RuntimeAffinity.LLAMACPP: FakeLlamaCppRuntime()},
    )
    try:
        baseline_app = create_app(temp_settings, services=baseline_services)
        with TestClient(baseline_app) as client:
            scan_response = client.post("/v1/models/scan", json={})
            manifests = scan_response.json()["manifests"]
            gguf_model_id = next(manifest["model_id"] for manifest in manifests if manifest["format_type"] == "gguf")
            baseline_capability = client.get(f"/v1/models/{gguf_model_id}/capabilities").json()
    finally:
        baseline_services.close()

    validation_manifest_path = temp_settings.data_dir / "validation" / "linux-validation.json"
    _write_external_validation_manifest(
        validation_manifest_path,
        capability_report=baseline_capability,
        system="Linux",
        machine="x86_64",
    )

    validated_settings = temp_settings.with_updates(validation_manifest_paths=(validation_manifest_path,))
    validated_services = bootstrap_services(
        validated_settings,
        runtime_overrides={RuntimeAffinity.LLAMACPP: FakeLlamaCppRuntime()},
    )
    try:
        app = create_app(validated_settings, services=validated_services)
        with TestClient(app) as client:
            scan_response = client.post("/v1/models/scan", json={})
            manifests = scan_response.json()["manifests"]
            gguf_model_id = next(manifest["model_id"] for manifest in manifests if manifest["format_type"] == "gguf")
            capability_response = client.get(f"/v1/models/{gguf_model_id}/capabilities")
            missing_response = client.get("/v1/models/missing-model/capabilities")
    finally:
        validated_services.close()

    assert capability_response.status_code == 200
    payload = capability_response.json()
    assert payload["model_id"] == gguf_model_id
    assert payload["host_platform"]["system"]
    assert any(
        item["runtime_name"] == "fake_llamacpp"
        and item["available"] is True
        and item["readiness_state"] == "ready"
        for item in payload["runtime_candidates"]
    )
    assert any(
        item["capability"] == "chat"
        and item["supported"] is True
        and item["readiness_state"] == "ready"
        for item in payload["capabilities"]
    )
    measured = {item["category"]: item for item in payload["measured_capabilities"]}
    assert measured["constrained_decoding"]["status"] == "rejected"
    linux_target = next(item for item in payload["target_platforms"] if item["system"] == "Linux" and item["machine"] == "x86_64")
    assert linux_target["readiness_state"] == "verified_external"
    assert linux_target["verification_method"] == "external_release_manifest"
    assert linux_target["validation_manifest_count"] == 1

    assert missing_response.status_code == 404
    assert missing_response.json()["error"]["code"] == "model_not_found"


def test_skill_and_tool_catalog_endpoints_and_execution(
    temp_settings,
    contract_transform_payload: dict[str, object],
) -> None:
    app = create_app(temp_settings)

    with TestClient(app) as client:
        skills_response = client.get("/v1/skills")
        skill_response = client.get("/v1/skills/document_comparison")
        tools_response = client.get("/v1/tools")
        tool_response = client.get("/v1/tools/documents.transform")
        execute_response = client.post(
            "/v1/tools/execute",
            json={"tool": "documents.transform", "input": contract_transform_payload},
        )

    assert skills_response.status_code == 200
    assert skills_response.json()["count"] >= 4
    assert any(item["name"] == "document_comparison" for item in skills_response.json()["items"])
    assert skill_response.status_code == 200
    assert skill_response.json()["tool_name"] == "documents.transform"

    assert tools_response.status_code == 200
    assert tools_response.json()["count"] >= 3
    assert any(item["name"] == "documents.ingest" for item in tools_response.json()["items"])
    assert tool_response.status_code == 200
    assert tool_response.json()["required_authorization"] == "document_transform"

    assert execute_response.status_code == 200
    payload = execute_response.json()
    assert payload["request_id"]
    assert payload["tool"] == "documents.transform"
    assert payload["trace"]["required_authorization"] == "document_transform"
    assert payload["trace"]["details"]["sandboxed"] is True
    assert payload["result"]["skill"] == "contract_text_replacement"
    assert payload["result"]["content_base64"]


def test_tool_execution_endpoint_emits_lifecycle_events(
    app_with_fake_runtime,
    contract_transform_payload: dict[str, object],
) -> None:
    with TestClient(app_with_fake_runtime) as client:
        with client.websocket_connect("/v1/events") as websocket:
            execute_response = client.post(
                "/v1/tools/execute",
                json={"tool": "documents.transform", "input": contract_transform_payload},
            )
            assert execute_response.status_code == 200
            request_id = execute_response.json()["request_id"]
            events = [websocket.receive_json() for _ in range(10)]

    assert [event["type"] for event in events] == [
        "tool.pending",
        "tool.started",
        "document.transform.started",
        "operation.progress",
        "document.render.started",
        "operation.progress",
        "operation.progress",
        "document.render.completed",
        "document.transform.completed",
        "tool.finished",
    ]
    assert all(event["payload"]["request_id"] == request_id for event in events)


def test_tool_and_document_endpoints_replay_idempotent_operations(
    temp_settings,
    contract_transform_payload: dict[str, object],
    sample_document_payload: dict[str, object],
) -> None:
    app = create_app(temp_settings)

    with TestClient(app) as client:
        first_generate = client.post(
            "/v1/documents/generate",
            json={
                "output_format": "csv",
                "file_name": "report.csv",
                "document": sample_document_payload,
                "idempotency_key": "doc-generate-1",
            },
        )
        second_generate = client.post(
            "/v1/documents/generate",
            json={
                "output_format": "csv",
                "file_name": "report.csv",
                "document": sample_document_payload,
                "idempotency_key": "doc-generate-1",
            },
        )
        first_tool = client.post(
            "/v1/tools/execute",
            json={
                "tool": "documents.transform",
                "input": {**contract_transform_payload, "idempotency_key": "tool-transform-1"},
            },
        )
        second_tool = client.post(
            "/v1/tools/execute",
            json={
                "tool": "documents.transform",
                "input": {**contract_transform_payload, "idempotency_key": "tool-transform-1"},
            },
        )

    assert first_generate.status_code == 200
    assert second_generate.status_code == 200
    assert first_generate.json()["idempotent_replay"] is False
    assert second_generate.json()["idempotent_replay"] is True
    assert second_generate.json()["request_id"] == first_generate.json()["request_id"]
    assert second_generate.json()["idempotency_key"] == "doc-generate-1"
    assert first_generate.json()["metadata"]["result_origin"] == "tool_execution"
    assert second_generate.json()["metadata"]["result_origin"] == "idempotent_replay"
    assert first_generate.json()["metadata"]["routing"]["reason"] == "documents.generate"

    assert first_tool.status_code == 200
    assert second_tool.status_code == 200
    assert first_tool.json()["idempotent_replay"] is False
    assert second_tool.json()["idempotent_replay"] is True
    assert second_tool.json()["request_id"] == first_tool.json()["request_id"]
    assert second_tool.json()["idempotency_key"] == "tool-transform-1"


def test_document_generation_endpoint_rejects_conflicting_idempotency_keys(
    temp_settings,
    sample_document_payload: dict[str, object],
) -> None:
    app = create_app(temp_settings)

    with TestClient(app) as client:
        first_response = client.post(
            "/v1/documents/generate",
            json={
                "output_format": "csv",
                "file_name": "report.csv",
                "document": sample_document_payload,
                "idempotency_key": "doc-conflict-1",
            },
        )
        conflicting_response = client.post(
            "/v1/documents/generate",
            json={
                "output_format": "markdown",
                "file_name": "report.md",
                "document": sample_document_payload,
                "idempotency_key": "doc-conflict-1",
            },
        )

    assert first_response.status_code == 200
    assert conflicting_response.status_code == 409
    assert conflicting_response.json()["error"]["code"] == "idempotency_conflict"
    assert conflicting_response.json()["error"]["details"]["idempotency_key"] == "doc-conflict-1"
    assert conflicting_response.json()["error"]["details"]["fallback_guidance"]


def test_document_generation_endpoint_supports_all_renderers(
    temp_settings,
    sample_document_payload: dict[str, object],
) -> None:
    app = create_app(temp_settings)

    with TestClient(app) as client:
        responses = {
            output_format: client.post(
                "/v1/documents/generate",
                json={
                    "output_format": output_format,
                    "file_name": {
                        "text": "report.txt",
                        "markdown": "report.md",
                        "json": "report.json",
                    }.get(output_format, f"report.{output_format}"),
                    "document": sample_document_payload,
                },
            )
            for output_format in ("text", "markdown", "json", "csv", "docx", "pdf")
        }
        responses["xlsx"] = client.post(
            "/v1/documents/generate",
            json={
                "output_format": "xlsx",
                "file_name": "report.xlsx",
                "document": sample_document_payload,
            },
        )

    assert responses["text"].status_code == 200
    assert base64.b64decode(responses["text"].json()["content_base64"]).decode("utf-8").startswith("Quarterly Operations Summary")
    assert responses["markdown"].status_code == 200
    assert base64.b64decode(responses["markdown"].json()["content_base64"]).decode("utf-8").startswith("# Quarterly Operations Summary")
    assert responses["json"].status_code == 200
    assert '"title": "Quarterly Operations Summary"' in base64.b64decode(responses["json"].json()["content_base64"]).decode("utf-8")
    assert responses["csv"].status_code == 200
    assert responses["csv"].json()["request_id"]
    assert base64.b64decode(responses["csv"].json()["content_base64"]).decode("utf-8").startswith("Category,Amount")
    assert base64.b64decode(responses["docx"].json()["content_base64"]).startswith(b"PK")
    assert base64.b64decode(responses["pdf"].json()["content_base64"]).startswith(b"%PDF")
    workbook = load_workbook(BytesIO(base64.b64decode(responses["xlsx"].json()["content_base64"])))
    assert workbook["Budget"]["A4"].value == "Category"


def test_document_transform_endpoint_supports_contract_receipt_and_file_template_skills(
    temp_settings,
    contract_transform_payload: dict[str, object],
    receipt_transform_payload: dict[str, object],
    branded_document_template_payload: dict[str, object],
    ocr_assisted_extraction_payload: dict[str, object],
    file_template_transform_payload: dict[str, object],
    document_compare_transform_payload: dict[str, object],
    meeting_transcript_notes_payload: dict[str, object],
    long_document_memo_payload: dict[str, object],
    speech_transcript_cleanup_payload: dict[str, object],
) -> None:
    app = create_app(temp_settings)

    with TestClient(app) as client:
        contract_response = client.post("/v1/documents/transform", json=contract_transform_payload)
        receipt_response = client.post("/v1/documents/transform", json=receipt_transform_payload)
        branded_response = client.post("/v1/documents/transform", json=branded_document_template_payload)
        ocr_response = client.post("/v1/documents/transform", json=ocr_assisted_extraction_payload)
        template_response = client.post("/v1/documents/transform", json=file_template_transform_payload)
        compare_response = client.post("/v1/documents/transform", json=document_compare_transform_payload)
        meeting_response = client.post("/v1/documents/transform", json=meeting_transcript_notes_payload)
        memo_response = client.post("/v1/documents/transform", json=long_document_memo_payload)
        cleanup_response = client.post("/v1/documents/transform", json=speech_transcript_cleanup_payload)

    assert contract_response.status_code == 200
    assert contract_response.json()["request_id"]
    assert contract_response.json()["skill"] == "contract_text_replacement"
    assert base64.b64decode(contract_response.json()["content_base64"]).startswith(b"PK")

    assert receipt_response.status_code == 200
    assert receipt_response.json()["skill"] == "receipt_extraction"
    assert base64.b64decode(receipt_response.json()["content_base64"]).decode("utf-8").startswith(
        "Description,Quantity,Unit Price,Line Total",
    )

    assert branded_response.status_code == 200
    assert branded_response.json()["skill"] == "branded_document_template"
    branded_document = json.loads(base64.b64decode(branded_response.json()["content_base64"]).decode("utf-8"))
    assert branded_document["metadata"]["skill"] == "branded_document_template"
    assert branded_document["header"]["center"] == "Product Launch Brief"
    assert branded_document["sections"][0]["heading"] == "Overview"
    image_blocks = [
        block
        for section in branded_document["sections"]
        for block in section["blocks"]
        if block["type"] == "image"
    ]
    assert len(image_blocks) == 2

    assert ocr_response.status_code == 200
    assert ocr_response.json()["skill"] == "ocr_assisted_extraction"
    ocr_markdown = base64.b64decode(ocr_response.json()["content_base64"]).decode("utf-8")
    assert ocr_markdown.startswith("# Scanned Invoice Extraction")
    assert "Invoice Date | 2026-04-12 | extracted" in ocr_markdown
    assert "Payment Terms |  | missing" in ocr_markdown

    assert template_response.status_code == 200
    assert template_response.json()["skill"] == "file_template"
    workbook = load_workbook(BytesIO(base64.b64decode(template_response.json()["content_base64"])))
    assert workbook["Overview"]["B9"].value == "On Track"

    assert compare_response.status_code == 200
    assert compare_response.json()["skill"] == "document_comparison"
    comparison_workbook = load_workbook(BytesIO(base64.b64decode(compare_response.json()["content_base64"])))
    assert comparison_workbook["Overview"]["B8"].value == "Baseline Agreement"
    assert comparison_workbook["Only in Revised Agreement"]["A3"].value == "- Updated pricing schedule applies."

    assert meeting_response.status_code == 200
    assert meeting_response.json()["skill"] == "meeting_transcript_notes"
    meeting_markdown = base64.b64decode(meeting_response.json()["content_base64"]).decode("utf-8")
    assert meeting_markdown.startswith("# Project Kickoff Notes")
    assert "## Decisions" in meeting_markdown
    assert "Priya: Decision: postpone OCR improvements to the next sprint." in meeting_markdown

    assert memo_response.status_code == 200
    assert memo_response.json()["skill"] == "long_document_memo"
    memo_markdown = base64.b64decode(memo_response.json()["content_base64"]).decode("utf-8")
    assert memo_markdown.startswith("# Platform Readiness Memo")
    assert "## Key Highlights" in memo_markdown
    assert "What operator guidance is still missing for Linux and Windows rollouts?" in memo_markdown

    assert cleanup_response.status_code == 200
    assert cleanup_response.json()["skill"] == "speech_transcript_cleanup"
    cleanup_markdown = base64.b64decode(cleanup_response.json()["content_base64"]).decode("utf-8")
    assert cleanup_markdown.startswith("# Customer Call Cleanup")
    assert "Agent | We can confirm the friday deployment window." in cleanup_markdown
    assert "Unknown | Follow up with legal for the final notice." in cleanup_markdown


def test_document_ingest_endpoint_supports_pdf_and_image_bundle(
    temp_settings,
    sample_ingest_sources,
) -> None:
    app = create_app(temp_settings)

    with TestClient(app) as client:
        pdf_response = client.post("/v1/documents/ingest", json={"paths": [str(sample_ingest_sources["pdf"])]})
        image_response = client.post("/v1/documents/ingest", json={"paths": [str(sample_ingest_sources["image_bundle"])]})

    assert pdf_response.status_code == 200
    assert pdf_response.json()["request_id"]
    assert pdf_response.json()["sources"][0]["source_type"] == "pdf"
    assert pdf_response.json()["sources"][0]["source_label"] == "sample.pdf"
    assert pdf_response.json()["sources"][0]["media_type"] == "application/pdf"
    assert pdf_response.json()["chunks"]
    assert pdf_response.json()["chunks"][0]["source_id"] == pdf_response.json()["sources"][0]["source_id"]
    assert pdf_response.json()["chunks"][0]["section_label"].endswith("Page 1")
    assert pdf_response.json()["chunks"][0]["metadata"]["page_number"] == 1
    assert any(
        block["type"] == "paragraph" and "Operations remained on track" in block["text"]
        for section in pdf_response.json()["document"]["sections"]
        for block in section["blocks"]
        if block["type"] == "paragraph"
    )

    assert image_response.status_code == 200
    assert image_response.json()["request_id"]
    assert image_response.json()["sources"][0]["source_type"] == "image_bundle"
    assert image_response.json()["document"]["metadata"]["chunk_count"] == len(image_response.json()["chunks"])
    assert any(
        block["type"] == "image"
        for block in image_response.json()["document"]["sections"][0]["blocks"]
    )


def test_document_ingest_endpoint_supports_text_and_markdown(
    temp_settings,
    sample_ingest_sources,
) -> None:
    app = create_app(temp_settings)

    with TestClient(app) as client:
        text_response = client.post("/v1/documents/ingest", json={"paths": [str(sample_ingest_sources["text"])]})
        markdown_response = client.post("/v1/documents/ingest", json={"paths": [str(sample_ingest_sources["markdown"])]})

    assert text_response.status_code == 200
    assert text_response.json()["sources"][0]["source_type"] == "text"
    assert any(
        block["type"] == "paragraph" and "Escalate Linux host audit follow-up next." in block["text"]
        for section in text_response.json()["document"]["sections"]
        for block in section["blocks"]
        if block["type"] == "paragraph"
    )

    assert markdown_response.status_code == 200
    assert markdown_response.json()["sources"][0]["source_type"] == "markdown"
    assert markdown_response.json()["sources"][0]["source_label"] == "sample.md"
    assert markdown_response.json()["document"]["sections"][0]["metadata"]["section_label"] == "sample.md / Quarterly operations summary"
    assert any(
        block["type"] == "list" and block["items"] == ["Confirm Linux validation host booking", "Refresh milestone tracker"]
        for section in markdown_response.json()["document"]["sections"]
        for block in section["blocks"]
        if block["type"] == "list"
    )
    assert markdown_response.json()["sources"][0]["metadata"]["list_count"] == 2


def test_document_ingest_endpoint_emits_lifecycle_events(
    app_with_fake_runtime,
    sample_ingest_sources,
) -> None:
    with TestClient(app_with_fake_runtime) as client:
        with client.websocket_connect("/v1/events") as websocket:
            ingest_response = client.post("/v1/documents/ingest", json={"paths": [str(sample_ingest_sources["pdf"])]})
            assert ingest_response.status_code == 200
            request_id = ingest_response.json()["request_id"]
            events = [websocket.receive_json() for _ in range(3)]

    assert [event["type"] for event in events] == [
        "document.parse.started",
        "operation.progress",
        "document.parse.completed",
    ]
    assert all(event["payload"]["request_id"] == request_id for event in events)


def test_document_generation_endpoint_emits_render_events(
    app_with_fake_runtime,
    sample_document_payload: dict[str, object],
) -> None:
    with TestClient(app_with_fake_runtime) as client:
        with client.websocket_connect("/v1/events") as websocket:
            response = client.post(
                "/v1/documents/generate",
                json={
                    "output_format": "csv",
                    "file_name": "report.csv",
                    "document": sample_document_payload,
                },
            )
            assert response.status_code == 200
            request_id = response.json()["request_id"]
            events = [websocket.receive_json() for _ in range(4)]

    assert [event["type"] for event in events] == [
        "document.render.started",
        "operation.progress",
        "operation.progress",
        "document.render.completed",
    ]
    assert all(event["payload"]["request_id"] == request_id for event in events)
    assert all(event["capability"] == "documents" for event in events)
    assert events[0]["operation"] == "document.render"
    assert events[0]["status"] == "started"
    assert events[-1]["status"] == "completed"


def test_multimodal_endpoints_support_embeddings_rerank_and_audio(
    app_with_fake_multimodal_runtime,
    sample_audio_bytes: bytes,
) -> None:
    with TestClient(app_with_fake_multimodal_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        manifests = scan_response.json()["manifests"]
        embedding_model_id = next(
            manifest["model_id"]
            for manifest in manifests
            if manifest["display_name"] == "e5-small-embed-mlx"
        )
        rerank_model_id = next(
            manifest["model_id"]
            for manifest in manifests
            if manifest["display_name"] == "bge-reranker-base-mlx"
        )
        audio_model_id = next(
            manifest["model_id"]
            for manifest in manifests
            if manifest["display_name"] == "whisper-mini-audio"
        )

        embeddings_response = client.post(
            "/v1/embeddings",
            json={"model": embedding_model_id, "input": ["alpha", "beta"]},
        )
        rerank_response = client.post(
            "/v1/rerank",
            json={
                "model": rerank_model_id,
                "query": "local model",
                "documents": ["local model routing", "remote api", "local document model"],
                "top_n": 2,
            },
        )
        retrieval_response = client.post(
            "/v1/retrieval/context",
            json={
                "embedding_model": embedding_model_id,
                "rerank_model": rerank_model_id,
                "query": "local model",
                "candidate_sources": [
                    {
                        "source_id": "source-1",
                        "path": "/tmp/source.md",
                        "source_type": "markdown",
                        "source_name": "source.md",
                        "source_label": "source.md",
                    }
                ],
                "candidate_chunks": [
                    {
                        "chunk_id": "chunk-1",
                        "text": "local model routing",
                        "source_id": "source-1",
                        "section_id": "section-1",
                        "source_label": "source.md",
                        "section_label": "source.md / Section 1",
                        "source_name": "source.md",
                        "source_path": "/tmp/source.md",
                        "source_type": "markdown",
                    },
                    {
                        "chunk_id": "chunk-2",
                        "text": "remote api",
                        "source_id": "source-1",
                        "section_id": "section-2",
                        "source_label": "source.md",
                        "section_label": "source.md / Section 2",
                        "source_name": "source.md",
                        "source_path": "/tmp/source.md",
                        "source_type": "markdown",
                    },
                    {
                        "chunk_id": "chunk-3",
                        "text": "local document model",
                        "source_id": "source-1",
                        "section_id": "section-3",
                        "source_label": "source.md",
                        "section_label": "source.md / Section 3",
                        "source_name": "source.md",
                        "source_path": "/tmp/source.md",
                        "source_type": "markdown",
                    },
                ],
                "top_k": 2,
            },
        )
        transcription_response = client.post(
            "/v1/audio/transcriptions",
            json={
                "model": audio_model_id,
                "audio_base64": base64.b64encode(sample_audio_bytes).decode("ascii"),
                "file_name": "sample.wav",
            },
        )
        speech_response = client.post(
            "/v1/audio/speech",
            json={"model": audio_model_id, "input": "Hello from LewLM", "voice": "alloy", "format": "wav"},
        )

    assert scan_response.status_code == 200
    assert scan_response.json()["discovered_count"] == 3

    assert embeddings_response.status_code == 200
    assert embeddings_response.json()["request_id"]
    assert embeddings_response.json()["created"] > 0
    assert embeddings_response.json()["object"] == "list"
    assert embeddings_response.json()["routing"]["model_id"] == embedding_model_id
    assert embeddings_response.json()["metadata"]["model"]["resolved_model_id"] == embedding_model_id
    assert embeddings_response.json()["metadata"]["model"]["runtime_name"] == embeddings_response.json()["routing"]["runtime_name"]
    assert embeddings_response.json()["metadata"]["timing"]["total_milliseconds"] >= 0
    assert len(embeddings_response.json()["data"]) == 2
    assert len(embeddings_response.json()["data"][0]["embedding"]) == 8

    assert rerank_response.status_code == 200
    assert rerank_response.json()["request_id"]
    assert rerank_response.json()["created"] > 0
    assert rerank_response.json()["routing"]["model_id"] == rerank_model_id
    assert rerank_response.json()["metadata"]["model"]["resolved_model_id"] == rerank_model_id
    assert rerank_response.json()["metadata"]["timing"]["total_milliseconds"] >= 0
    assert [item["index"] for item in rerank_response.json()["results"]] == [0, 2]

    assert retrieval_response.status_code == 200
    assert retrieval_response.json()["request_id"]
    assert retrieval_response.json()["created"] > 0
    assert retrieval_response.json()["strategy"] == "hybrid"
    assert retrieval_response.json()["candidate_count"] == 3
    assert retrieval_response.json()["returned_count"] == 2
    assert [item["chunk"]["chunk_id"] for item in retrieval_response.json()["items"]] == ["chunk-3", "chunk-1"]
    assert retrieval_response.json()["embedding_stage"]["model"] == embedding_model_id
    assert retrieval_response.json()["rerank_stage"]["model"] == rerank_model_id
    assert retrieval_response.json()["metadata"]["model"]["resolved_model_id"] == rerank_model_id
    assert retrieval_response.json()["metadata"]["timing"]["total_milliseconds"] >= 0

    assert transcription_response.status_code == 200
    assert transcription_response.json()["request_id"]
    assert transcription_response.json()["created"] > 0
    assert transcription_response.json()["routing"]["model_id"] == audio_model_id
    assert transcription_response.json()["metadata"]["model"]["resolved_model_id"] == audio_model_id
    assert transcription_response.json()["metadata"]["timing"]["total_milliseconds"] >= 0
    assert transcription_response.json()["text"].startswith("Transcribed sample.wav")

    assert speech_response.status_code == 200
    assert speech_response.json()["request_id"]
    assert speech_response.json()["created"] > 0
    assert speech_response.json()["routing"]["model_id"] == audio_model_id
    assert speech_response.json()["metadata"]["model"]["resolved_model_id"] == audio_model_id
    assert speech_response.json()["metadata"]["timing"]["total_milliseconds"] >= 0
    assert speech_response.json()["media_type"] == "audio/wav"
    assert speech_response.json()["content_type"] == "audio/wav"
    assert base64.b64decode(speech_response.json()["audio_base64"]).startswith(b"RIFF")


def test_audio_transcription_endpoint_emits_lifecycle_events(
    app_with_fake_multimodal_runtime,
    sample_audio_bytes: bytes,
) -> None:
    with TestClient(app_with_fake_multimodal_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        audio_model_id = next(
            manifest["model_id"]
            for manifest in scan_response.json()["manifests"]
            if manifest["display_name"] == "whisper-mini-audio"
        )
        with client.websocket_connect("/v1/events") as websocket:
            response = client.post(
                "/v1/audio/transcriptions",
                json={
                    "model": audio_model_id,
                    "audio_base64": base64.b64encode(sample_audio_bytes).decode("ascii"),
                    "file_name": "events.wav",
                },
            )
            assert response.status_code == 200
            request_id = response.json()["request_id"]
            events = [websocket.receive_json() for _ in range(8)]

    assert [event["type"] for event in events] == [
        "request.accepted",
        "audio.transcription.started",
        "operation.progress",
        "model.loading",
        "model.loaded",
        "operation.progress",
        "audio.transcription.completed",
        "request.completed",
    ]
    assert all(event["payload"]["request_id"] == request_id for event in events)
    assert all(event["request_id"] == request_id for event in events)
    assert all(event["capability"] == "audio_transcription" for event in events)
    assert events[0]["operation"] == "audio.transcription"
    assert events[0]["status"] == "accepted"
    assert events[-1]["status"] == "completed"
    assert events[2]["payload"]["operation"] == "audio.transcription"
    assert events[5]["payload"]["segment_count"] == 1


def test_audio_transcription_endpoint_emits_chunk_events_for_long_audio(
    app_with_fake_multimodal_runtime,
    long_sample_audio_bytes: bytes,
) -> None:
    with TestClient(app_with_fake_multimodal_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        audio_model_id = next(
            manifest["model_id"]
            for manifest in scan_response.json()["manifests"]
            if manifest["display_name"] == "whisper-mini-audio"
        )
        with client.websocket_connect("/v1/events") as websocket:
            response = client.post(
                "/v1/audio/transcriptions",
                json={
                    "model": audio_model_id,
                    "audio_base64": base64.b64encode(long_sample_audio_bytes).decode("ascii"),
                    "file_name": "long-events.wav",
                },
            )
            assert response.status_code == 200
            request_id = response.json()["request_id"]
            events = [websocket.receive_json() for _ in range(14)]

    assert [event["type"] for event in events] == [
        "request.accepted",
        "audio.transcription.started",
        "operation.progress",
        "model.loading",
        "model.loaded",
        "operation.progress",
        "audio.chunk",
        "operation.progress",
        "operation.progress",
        "audio.chunk",
        "operation.progress",
        "operation.progress",
        "audio.transcription.completed",
        "request.completed",
    ]
    assert all(event["payload"]["request_id"] == request_id for event in events)
    assert events[1]["payload"]["chunk_count"] == 2
    assert events[2]["payload"]["stage"] == "chunks_planned"
    assert events[2]["payload"]["chunk_count"] == 2
    assert events[5]["payload"]["stage"] == "chunk_started"
    assert events[6]["payload"]["chunk_index"] == 1
    assert events[7]["payload"]["stage"] == "chunk_processed"
    assert events[8]["payload"]["stage"] == "chunk_started"
    assert events[9]["payload"]["chunk_index"] == 2
    assert events[10]["payload"]["stage"] == "chunk_processed"
    assert events[11]["payload"]["stage"] == "segments_ready"
    assert events[11]["payload"]["segment_count"] == 2
    assert events[-2]["type"] == "audio.transcription.completed"
    assert events[-1]["type"] == "request.completed"


def test_audio_speech_endpoint_emits_lifecycle_events(
    app_with_fake_multimodal_runtime,
) -> None:
    with TestClient(app_with_fake_multimodal_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        audio_model_id = next(
            manifest["model_id"]
            for manifest in scan_response.json()["manifests"]
            if manifest["display_name"] == "whisper-mini-audio"
        )
        with client.websocket_connect("/v1/events") as websocket:
            response = client.post(
                "/v1/audio/speech",
                json={"model": audio_model_id, "input": "Ship the next milestone", "voice": "alloy", "format": "wav"},
            )
            assert response.status_code == 200
            request_id = response.json()["request_id"]
            events = [websocket.receive_json() for _ in range(8)]

    assert [event["type"] for event in events] == [
        "request.accepted",
        "audio.speech.started",
        "operation.progress",
        "model.loading",
        "model.loaded",
        "operation.progress",
        "audio.speech.completed",
        "request.completed",
    ]
    assert all(event["payload"]["request_id"] == request_id for event in events)
    assert events[2]["payload"]["operation"] == "audio.speech"
    assert events[5]["payload"]["audio_output_bytes"] > 0


def test_chat_completion_endpoint_emits_reasoning_safe_progress_events(
    app_with_fake_runtime,
) -> None:
    with TestClient(app_with_fake_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        gguf_model_id = next(
            manifest["model_id"]
            for manifest in scan_response.json()["manifests"]
            if manifest["format_type"] == "gguf"
        )
        with client.websocket_connect("/v1/events") as websocket:
            response = client.post(
                "/v1/chat/completions",
                json={
                    "model": gguf_model_id,
                    "messages": [{"role": "user", "content": "Show generation progress"}],
                },
            )
            assert response.status_code == 200
            request_id = response.json()["id"]
            events = [websocket.receive_json() for _ in range(9)]

    assert [event["type"] for event in events] == [
        "request.accepted",
        "operation.progress",
        "model.loading",
        "model.loaded",
        "operation.progress",
        "prefill.started",
        "operation.progress",
        "operation.progress",
        "request.completed",
    ]
    assert all(event["payload"]["request_id"] == request_id for event in events if "payload" in event)
    assert [events[1]["payload"]["stage"], events[4]["payload"]["stage"], events[6]["payload"]["stage"], events[7]["payload"]["stage"]] == [
        "prompt_compiled",
        "model_ready",
        "response_generating",
        "response_ready",
    ]
    assert all(events[index]["payload"]["reasoning_exposed"] is False for index in (1, 4, 6, 7))
    assert all(events[index]["payload"]["reasoning_visibility"] == "hidden" for index in (1, 4, 6, 7))
    assert events[0]["payload"]["serving"]["phase"] == "admitted"
    assert events[5]["payload"]["serving"]["phase"] == "prefill"
    assert events[-1]["payload"]["serving"]["phase"] == "completed"
    assert events[-1]["payload"]["serving"]["runtime_adapter"]["kind"] in {
        "backend_native_batch",
        "request_scoped",
    }


def test_chat_completion_endpoint_emits_speculation_summary_events(
    temp_settings,
    sample_models_root: Path,
) -> None:
    draft_dir = temp_settings.models_dir[0] / "qwen2.5-0.5b-instruct-draft-mlx"
    draft_dir.mkdir(parents=True)
    (draft_dir / "config.json").write_text(
        json.dumps({"model_type": "qwen2", "max_position_embeddings": 32768}),
        encoding="utf-8",
    )
    (draft_dir / "weights.safetensors").write_bytes(b"draft-weights")
    (draft_dir / "tokenizer.json").write_text("{}", encoding="utf-8")

    class _SpeculationEventRuntime(ManagedTextRuntime):
        name = "fake_mlx_speculation_events"
        affinity = RuntimeAffinity.MLX_TEXT
        supported_formats = (ModelFormat.MLX,)
        supported_modalities = (ModelModality.TEXT,)
        supported_capabilities = frozenset({CapabilityName.CHAT, CapabilityName.STREAMING})

        def _check_environment(self) -> tuple[bool, str | None]:
            return True, None

        def performance_feature_snapshot(self) -> dict[str, object]:
            return {
                "speculative_decoding": {
                    "supported": True,
                    "active": False,
                    "modes": ["draft_model"],
                    "reason": "Fake MLX runtime exposes LewLM-owned draft verification for event tests.",
                    "metrics": {},
                },
            }

        async def _load_model(self, manifest: ModelManifest) -> None:
            return None

        async def _unload_model(self, model_id: str) -> None:
            return None

        async def _generate(self, request: GenerateRequest) -> GenerateResponse:
            usage = {
                "prompt_tokens": len(request.messages),
                "completion_tokens": 2,
                "total_tokens": len(request.messages) + 2,
            }
            if request.speculation is not None:
                request.metadata["speculation_execution_path"] = "lewlm_controller"
                request.metadata["speculation_runtime"] = {
                    "execution_path": "lewlm_controller",
                    "controller": "draft_verify",
                    "drafted_tokens": 2,
                    "accepted_tokens": 2,
                    "verified_tokens": 2,
                    "rejected_tokens": 0,
                    "rollback_tokens": 0,
                    "fallback_count": 0,
                }
                usage.update(
                    {
                        "drafted_tokens": 2,
                        "accepted_tokens": 2,
                        "verified_tokens": 2,
                        "rejected_tokens": 0,
                        "rollback_tokens": 0,
                    },
                )
            return GenerateResponse(
                model_id=request.model_id,
                output_text="OK",
                finish_reason="stop",
                usage=usage,
            )

        async def _stream_generate(self, request: GenerateRequest):
            yield "OK"

        def _tokenize(self, text: str) -> list[int]:
            return list(text.encode("utf-8"))

        def _detokenize(self, tokens) -> str:
            return bytes(tokens).decode("utf-8")

    runtime = _SpeculationEventRuntime()
    settings = temp_settings.with_updates(
        speculative_decoding_enabled=True,
        speculative_decoding_num_draft_tokens=2,
    )
    services = bootstrap_services(
        settings,
        runtime_overrides={RuntimeAffinity.MLX_TEXT: runtime},
    )
    try:
        app = create_app(settings, services=services)
        with TestClient(app) as client:
            scan_response = client.post("/v1/models/scan", json={})
            mlx_model_id = next(
                manifest["model_id"]
                for manifest in scan_response.json()["manifests"]
                if manifest["display_name"] == "qwen2.5-1.5b-instruct-mlx"
            )
            with client.websocket_connect("/v1/events") as websocket:
                response = client.post(
                    "/v1/chat/completions",
                    json={
                        "model": mlx_model_id,
                        "messages": [{"role": "user", "content": "Show speculation events"}],
                    },
                )
                assert response.status_code == 200
                request_id = response.json()["id"]
                events = [websocket.receive_json() for _ in range(11)]
    finally:
        services.close()

    assert [event["type"] for event in events] == [
        "request.accepted",
        "operation.progress",
        "model.loading",
        "model.loaded",
        "operation.progress",
        "prefill.started",
        "speculation.started",
        "operation.progress",
        "operation.progress",
        "speculation.accepted",
        "request.completed",
    ]
    assert all(event["payload"]["request_id"] == request_id for event in events if "payload" in event)
    assert events[6]["payload"]["mode"] == "draft_model"
    assert events[6]["payload"]["execution_path"] == "pending"
    assert events[9]["payload"]["execution_path"] == "lewlm_controller"
    assert events[9]["payload"]["accepted_tokens"] == 2
    assert events[9]["payload"]["fallback_count"] == 0


def test_responses_endpoint_emits_reasoning_safe_progress_events(
    app_with_fake_runtime,
) -> None:
    with TestClient(app_with_fake_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        gguf_model_id = next(
            manifest["model_id"]
            for manifest in scan_response.json()["manifests"]
            if manifest["format_type"] == "gguf"
        )
        with client.websocket_connect("/v1/events") as websocket:
            response = client.post(
                "/v1/responses",
                json={"model": gguf_model_id, "input": "Show response progress"},
            )
            assert response.status_code == 200
            request_id = response.json()["id"]
            events = [websocket.receive_json() for _ in range(9)]

    assert [event["type"] for event in events] == [
        "request.accepted",
        "operation.progress",
        "model.loading",
        "model.loaded",
        "operation.progress",
        "prefill.started",
        "operation.progress",
        "operation.progress",
        "request.completed",
    ]
    assert all(event["payload"]["request_id"] == request_id for event in events if "payload" in event)
    assert events[1]["payload"]["operation"] == "text.generation"
    assert events[7]["payload"]["stage"] == "response_ready"
    assert all(events[index]["payload"]["reasoning_visibility"] == "hidden" for index in (1, 4, 6, 7))


def test_chat_completion_reasoning_visibility_exposes_model_emitted_reasoning(
    app_with_fake_runtime,
) -> None:
    with TestClient(app_with_fake_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        gguf_model_id = next(
            manifest["model_id"]
            for manifest in scan_response.json()["manifests"]
            if manifest["format_type"] == "gguf"
        )
        response = client.post(
            "/v1/chat/completions",
            json={
                "model": gguf_model_id,
                "reasoning_visibility": "raw_model_emitted",
                "messages": [{"role": "user", "content": "[emit-reasoning] Explain the plan"}],
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["choices"][0]["message"]["content"] == "Echo: Explain the plan"
    assert payload["choices"][0]["message"]["reasoning"] == {
        "visibility": "raw_model_emitted",
        "available": True,
        "content": "Inspect the prompt before replying.",
        "summary": None,
    }


def test_chat_and_responses_return_structured_citations_from_known_context(
    app_with_fake_runtime,
) -> None:
    citation_context = {
        "sources": [
            {
                "source_id": "source-1",
                "path": "/tmp/source.md",
                "source_type": "markdown",
                "source_name": "source.md",
                "source_label": "Source One",
                "metadata": {},
            }
        ],
        "chunks": [
            {
                "chunk_id": "chunk-1",
                "text": "LewLM exposes a local-first backend package.",
                "source_id": "source-1",
                "section_id": "section-1",
                "source_label": "Source One",
                "section_label": "Source One / Summary",
                "metadata": {},
            }
        ],
    }

    with TestClient(app_with_fake_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        gguf_model_id = next(
            manifest["model_id"]
            for manifest in scan_response.json()["manifests"]
            if manifest["format_type"] == "gguf"
        )
        chat_response = client.post(
            "/v1/chat/completions",
            json={
                "model": gguf_model_id,
                "messages": [{"role": "user", "content": "Summarize LewLM with grounding"}],
                "citation_context": citation_context,
            },
        )
        responses_api = client.post(
            "/v1/responses",
            json={
                "model": gguf_model_id,
                "input": "Summarize LewLM with grounding",
                "citation_context": citation_context,
            },
        )

    assert chat_response.status_code == 200
    chat_payload = chat_response.json()
    assert chat_payload["choices"][0]["message"]["content"] == "Echo: Summarize LewLM with grounding"
    assert chat_payload["citations"] == [
        {
            "reference_id": "chunk-1",
            "source_id": "source-1",
            "chunk_id": "chunk-1",
            "section_id": "section-1",
            "source_label": "Source One",
            "section_label": "Source One / Summary",
        }
    ]

    assert responses_api.status_code == 200
    responses_payload = responses_api.json()
    assert responses_payload["output_text"] == "Echo: Summarize LewLM with grounding"
    assert responses_payload["citations"] == chat_payload["citations"]


def test_streaming_chat_returns_final_citation_package(
    app_with_fake_runtime,
) -> None:
    citation_context = {
        "sources": [
            {
                "source_id": "source-1",
                "path": "/tmp/source.md",
                "source_type": "markdown",
                "source_name": "source.md",
                "source_label": "Source One",
                "metadata": {},
            }
        ],
        "chunks": [
            {
                "chunk_id": "chunk-1",
                "text": "LewLM exposes a local-first backend package.",
                "source_id": "source-1",
                "section_id": "section-1",
                "source_label": "Source One",
                "section_label": "Source One / Summary",
                "metadata": {},
            }
        ],
    }

    with TestClient(app_with_fake_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        gguf_model_id = next(
            manifest["model_id"]
            for manifest in scan_response.json()["manifests"]
            if manifest["format_type"] == "gguf"
        )
        with client.stream(
            "POST",
            "/v1/chat/completions",
            json={
                "model": gguf_model_id,
                "messages": [{"role": "user", "content": "Stream LewLM with grounding"}],
                "citation_context": citation_context,
                "stream": True,
            },
        ) as stream_response:
            data_lines = [
                line.removeprefix("data: ")
                for line in stream_response.iter_lines()
                if line and line.startswith("data: ")
            ]

    assert data_lines[-1] == "[DONE]"
    stream_chunks = [json.loads(line) for line in data_lines[:-1]]
    streamed_text = "".join(
        chunk["choices"][0]["delta"].get("content", "")
        for chunk in stream_chunks
        if chunk["choices"][0]["delta"].get("content")
    )
    assert streamed_text == "Echo: Stream LewLM with grounding"
    assert stream_chunks[-1]["citations"] == [
        {
            "reference_id": "chunk-1",
            "source_id": "source-1",
            "chunk_id": "chunk-1",
            "section_id": "section-1",
            "source_label": "Source One",
            "section_label": "Source One / Summary",
        }
    ]


def test_chat_completion_errors_include_structured_fallback_guidance(
    app_with_fake_multimodal_runtime,
) -> None:
    with TestClient(app_with_fake_multimodal_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        audio_model_id = next(
            manifest["model_id"]
            for manifest in scan_response.json()["manifests"]
            if manifest["display_name"] == "whisper-mini-audio"
        )
        response = client.post(
            "/v1/chat/completions",
            json={
                "model": audio_model_id,
                "messages": [{"role": "user", "content": "Hello from the wrong modality"}],
            },
        )

    assert response.status_code == 400
    payload = response.json()["error"]
    assert payload["code"] == "routing_error"
    assert payload["details"]["requested_model_id"] == audio_model_id
    assert payload["details"]["required_modalities"] == ["text"]
    assert payload["details"]["fallback_guidance"]


def test_multimodal_endpoints_accept_multipart_uploads(
    temp_settings,
    app_with_fake_attachment_runtime,
    sample_attachment_sources,
    sample_audio_bytes: bytes,
) -> None:
    with TestClient(app_with_fake_attachment_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        manifests = scan_response.json()["manifests"]
        vision_model_id = next(
            manifest["model_id"]
            for manifest in manifests
            if manifest["display_name"] == "qwen2-vl-vision-mlx"
        )
        audio_model_id = next(
            manifest["model_id"]
            for manifest in manifests
            if manifest["display_name"] == "whisper-mini-audio"
        )

        chat_response = client.post(
            "/v1/chat/completions",
            data={
                "payload_json": json.dumps(
                    {
                        "messages": [
                            {
                                "role": "user",
                                "content": [
                                    {"type": "text", "text": "Review these uploaded files."},
                                    {"type": "input_image", "upload_name": "image_upload"},
                                    {"type": "input_file", "upload_name": "document_upload"},
                                    {"type": "input_audio", "upload_name": "audio_upload"},
                                ],
                            },
                        ],
                    },
                ),
            },
            files=[
                ("image_upload", ("receipt-front.png", sample_attachment_sources["image_one"].read_bytes(), "image/png")),
                (
                    "document_upload",
                    (
                        "sample.pdf",
                        sample_attachment_sources["pdf"].read_bytes(),
                        "application/pdf",
                    ),
                ),
                ("audio_upload", ("voice-note.wav", sample_audio_bytes, "audio/wav")),
            ],
        )
        response_api = client.post(
            "/v1/responses",
            data={
                "payload_json": json.dumps(
                    {
                        "input": [
                            {
                                "role": "user",
                                "content": [
                                    {"type": "input_text", "text": "Summarize this uploaded spreadsheet."},
                                    {"type": "input_file", "upload_name": "sheet_upload"},
                                ],
                            },
                        ],
                    },
                ),
            },
            files=[
                (
                    "sheet_upload",
                    (
                        "sample.xlsx",
                        sample_attachment_sources["xlsx"].read_bytes(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    ),
                ),
            ],
        )
        transcription_response = client.post(
            "/v1/audio/transcriptions",
            data={"model": audio_model_id},
            files=[("file", ("sample.wav", sample_audio_bytes, "audio/wav"))],
        )

    assert chat_response.status_code == 200
    assert chat_response.json()["model"] == vision_model_id
    assert "[Attached image: 0-receipt-front.png]" in chat_response.json()["choices"][0]["message"]["content"]
    assert "Transcribed 2-voice-note.wav" in chat_response.json()["choices"][0]["message"]["content"]

    assert response_api.status_code == 200
    assert "[Attached document: 0-sample.xlsx]" in response_api.json()["output_text"]

    assert transcription_response.status_code == 200
    assert transcription_response.json()["model"] == audio_model_id
    assert transcription_response.json()["text"].startswith("Transcribed sample.wav")
    assert _api_upload_workspaces(temp_settings) == []


def test_multipart_chat_request_cleans_up_workspace_after_duplicate_upload_failure(
    temp_settings,
    app_with_fake_attachment_runtime,
    sample_attachment_sources,
) -> None:
    with TestClient(app_with_fake_attachment_runtime) as client:
        response = client.post(
            "/v1/chat/completions",
            data={
                "payload_json": json.dumps(
                    {
                        "messages": [
                            {
                                "role": "user",
                                "content": [
                                    {"type": "text", "text": "Inspect duplicate uploads."},
                                    {"type": "input_image", "upload_name": "image_upload"},
                                ],
                            },
                        ],
                    },
                ),
            },
            files=[
                ("image_upload", ("receipt-front.png", sample_attachment_sources["image_one"].read_bytes(), "image/png")),
                ("image_upload", ("receipt-back.png", sample_attachment_sources["image_one"].read_bytes(), "image/png")),
            ],
        )

    assert response.status_code == 400
    assert response.json()["error"]["code"] == "configuration_error"
    assert response.json()["error"]["details"]["field_name"] == "image_upload"
    assert _api_upload_workspaces(temp_settings) == []


def test_multipart_chat_request_cleans_up_workspace_after_missing_upload_reference(
    temp_settings,
    app_with_fake_attachment_runtime,
    sample_attachment_sources,
) -> None:
    with TestClient(app_with_fake_attachment_runtime) as client:
        response = client.post(
            "/v1/chat/completions",
            data={
                "payload_json": json.dumps(
                    {
                        "messages": [
                            {
                                "role": "user",
                                "content": [
                                    {"type": "text", "text": "Reference a missing upload."},
                                    {"type": "input_image", "upload_name": "missing_upload"},
                                ],
                            },
                        ],
                    },
                ),
            },
            files=[
                ("provided_upload", ("receipt-front.png", sample_attachment_sources["image_one"].read_bytes(), "image/png")),
            ],
        )

    assert response.status_code == 400
    assert response.json()["error"]["code"] == "configuration_error"
    assert response.json()["error"]["details"]["upload_name"] == "missing_upload"
    assert _api_upload_workspaces(temp_settings) == []


def test_chat_and_responses_reject_out_of_scope_local_attachments(
    temp_settings,
    app_with_fake_attachment_runtime,
    sample_attachment_sources,
) -> None:
    outside_dir = temp_settings.data_dir.parent / "outside-attachments"
    outside_dir.mkdir(parents=True, exist_ok=True)
    outside_pdf = outside_dir / "outside.pdf"
    outside_audio = outside_dir / "outside.wav"
    outside_pdf.write_bytes(sample_attachment_sources["pdf"].read_bytes())
    outside_audio.write_bytes(sample_attachment_sources["audio"].read_bytes())

    with TestClient(app_with_fake_attachment_runtime) as client:
        chat_response = client.post(
            "/v1/chat/completions",
            json={
                "messages": [
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": "Read this forbidden document."},
                            {"type": "input_file", "path": str(outside_pdf)},
                        ],
                    },
                ],
            },
        )
        responses_response = client.post(
            "/v1/responses",
            json={
                "input": [
                    {
                        "role": "user",
                        "content": [
                            {"type": "input_text", "text": "Transcribe this forbidden audio."},
                            {"type": "input_audio", "path": str(outside_audio)},
                        ],
                    },
                ],
            },
        )

    assert chat_response.status_code == 403
    assert chat_response.json()["error"]["code"] == "file_access_error"
    assert chat_response.json()["error"]["details"]["path"] == str(outside_pdf.resolve(strict=False))

    assert responses_response.status_code == 403
    assert responses_response.json()["error"]["code"] == "file_access_error"
    assert responses_response.json()["error"]["details"]["path"] == str(outside_audio.resolve(strict=False))


def test_chat_and_responses_support_mixed_local_attachments(
    app_with_fake_attachment_runtime,
    sample_attachment_sources,
) -> None:
    with TestClient(app_with_fake_attachment_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        assert scan_response.status_code == 200
        vision_model_id = next(
            manifest["model_id"]
            for manifest in scan_response.json()["manifests"]
            if manifest["display_name"] == "qwen2-vl-vision-mlx"
        )

        chat_response = client.post(
            "/v1/chat/completions",
            json={
                "messages": [
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": "Summarize these local artifacts."},
                            {"type": "input_image", "path": str(sample_attachment_sources["image_one"])},
                            {"type": "input_file", "path": str(sample_attachment_sources["pdf"])},
                            {"type": "input_file", "path": str(sample_attachment_sources["text"])},
                            {"type": "input_audio", "path": str(sample_attachment_sources["audio"])},
                        ],
                    },
                ],
            },
        )
        response_api = client.post(
            "/v1/responses",
            json={
                "input": [
                    {
                        "role": "user",
                        "content": [
                            {"type": "input_text", "text": "Turn this spreadsheet into notes."},
                            {"type": "input_file", "path": str(sample_attachment_sources["xlsx"])},
                        ],
                    },
                ],
            },
        )

    assert chat_response.status_code == 200
    assert chat_response.json()["model"] == vision_model_id
    chat_text = chat_response.json()["choices"][0]["message"]["content"]
    assert chat_text.startswith("Vision echo: Summarize these local artifacts.")
    assert "[Attached image: receipt-front.png]" in chat_text
    assert "[Attached document: sample.pdf]" in chat_text
    assert "Attachment note: summarize the local progress update." in chat_text
    assert "Transcribed voice-note.wav" in chat_text
    assert "[images: receipt-front.png]" in chat_text

    assert response_api.status_code == 200
    output_text = response_api.json()["output_text"]
    assert "Turn this spreadsheet into notes." in output_text
    assert "[Attached document: sample.xlsx]" in output_text
    assert "Sheet Budget" in output_text


def test_chat_attachment_feature_cache_reuses_and_invalidates_local_artifacts(
    app_with_fake_attachment_runtime,
    sample_attachment_sources,
) -> None:
    with TestClient(app_with_fake_attachment_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        assert scan_response.status_code == 200
        vision_model_id = next(
            manifest["model_id"]
            for manifest in scan_response.json()["manifests"]
            if manifest["display_name"] == "qwen2-vl-vision-mlx"
        )

        def chat_payload() -> dict[str, object]:
            return {
                "model": vision_model_id,
                "messages": [
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": "Summarize these cached local artifacts."},
                            {"type": "input_image", "path": str(sample_attachment_sources["image_one"])},
                            {"type": "input_file", "path": str(sample_attachment_sources["text"])},
                        ],
                    },
                ],
            }

        first_response = client.post("/v1/chat/completions", json=chat_payload())
        second_response = client.post("/v1/chat/completions", json=chat_payload())
        sample_attachment_sources["text"].write_text(
            "Attachment note: cache invalidated after the local edit.",
            encoding="utf-8",
        )
        third_response = client.post("/v1/chat/completions", json=chat_payload())
        cache_response = client.get("/v1/cache/stats")
        runtime_response = client.get("/v1/runtime/stats")

    assert first_response.status_code == 200
    assert second_response.status_code == 200
    assert third_response.status_code == 200
    assert "Attachment note: summarize the local progress update." in first_response.json()["choices"][0]["message"]["content"]
    assert "Attachment note: summarize the local progress update." in second_response.json()["choices"][0]["message"]["content"]
    assert "Attachment note: cache invalidated after the local edit." in third_response.json()["choices"][0]["message"]["content"]

    cache_payload = cache_response.json()
    assert cache_payload["block_cache_count"] == 4
    assert cache_payload["multimodal_feature_count"] == 3
    assert cache_payload["multimodal_encoder_count"] == 1
    assert cache_payload["block_cache_bytes"] > 0
    assert cache_payload["multimodal_feature_bytes"] > 0
    assert cache_payload["multimodal_encoder_bytes"] > 0
    assert cache_payload["block_cache_hits"] == 3
    assert cache_payload["block_cache_misses"] == 3
    assert cache_payload["multimodal_feature_cache_hits"] == 3
    assert cache_payload["multimodal_feature_cache_misses"] == 3
    assert cache_payload["multimodal_encoder_cache_hits"] == 2
    assert cache_payload["multimodal_encoder_cache_misses"] == 1
    cache_features = {item["feature"]: item for item in cache_payload["performance_features"]}
    assert cache_features["block_disk_cache"]["supported"] is True
    assert cache_features["multimodal_feature_caching"]["supported"] is True
    assert cache_features["multimodal_encoder_caching"]["supported"] is True

    runtime_features = {item["feature"]: item for item in runtime_response.json()["performance_features"]}
    assert runtime_features["block_disk_cache"]["supported"] is True
    assert runtime_features["multimodal_feature_caching"]["supported"] is True
    assert runtime_features["multimodal_encoder_caching"]["supported"] is True


def test_session_endpoints_support_create_chat_export_import_and_messages(
    app_with_fake_runtime_session_enabled,
) -> None:
    with TestClient(app_with_fake_runtime_session_enabled) as client:
        scan_response = client.post("/v1/models/scan", json={})
        manifests = scan_response.json()["manifests"]
        gguf_model_id = next(manifest["model_id"] for manifest in manifests if manifest["format_type"] == "gguf")

        create_response = client.post("/v1/sessions", json={"title": "Milestone 9"})
        assert create_response.status_code == 200
        session_id = create_response.json()["session_id"]

        first_chat = client.post(
            "/v1/chat/completions",
            json={
                "model": gguf_model_id,
                "session_id": session_id,
                "messages": [{"role": "user", "content": "Capture the first note"}],
            },
        )
        second_response = client.post(
            "/v1/responses",
            json={
                "model": gguf_model_id,
                "session_id": session_id,
                "input": "Add the follow-up note",
            },
        )
        detail_response = client.get(f"/v1/sessions/{session_id}")
        messages_response = client.get(f"/v1/sessions/{session_id}/messages")
        export_response = client.get(f"/v1/sessions/{session_id}/export")
        import_response = client.post(
            "/v1/sessions/import",
            json={"bundle": export_response.json(), "title": "Imported milestone 9"},
        )
        list_response = client.get("/v1/sessions")

    assert first_chat.status_code == 200
    assert first_chat.json()["session_id"] == session_id
    assert second_response.status_code == 200
    assert second_response.json()["session_id"] == session_id

    assert detail_response.status_code == 200
    detail_payload = detail_response.json()
    assert detail_payload["turn_count"] == 2
    assert detail_payload["message_count"] == 4
    assert detail_payload["turns"][0]["input_messages"][0]["content"] == "Capture the first note"
    assert detail_payload["turns"][1]["response_message"]["content"] == "Echo: Add the follow-up note"
    assert "prefix_cache" in detail_payload["turns"][0]["metadata"]
    assert "prefix_cache" in detail_payload["turns"][1]["metadata"]

    assert messages_response.status_code == 200
    flattened_messages = messages_response.json()["messages"]
    assert [message["role"] for message in flattened_messages] == ["user", "assistant", "user", "assistant"]

    assert export_response.status_code == 200
    export_payload = export_response.json()
    assert export_payload["session"]["session_id"] == session_id
    assert len(export_payload["turns"]) == 2

    assert import_response.status_code == 200
    imported_payload = import_response.json()
    assert imported_payload["session_id"] != session_id
    assert imported_payload["title"] == "Imported milestone 9"
    assert imported_payload["turn_count"] == 2

    assert list_response.status_code == 200
    assert list_response.json()["count"] == 2


def test_session_context_policies_compact_history_during_chat(
    app_with_fake_runtime_session_enabled,
) -> None:
    with TestClient(app_with_fake_runtime_session_enabled) as client:
        scan_response = client.post("/v1/models/scan", json={})
        manifests = scan_response.json()["manifests"]
        gguf_model_id = next(manifest["model_id"] for manifest in manifests if manifest["format_type"] == "gguf")

        create_response = client.post(
            "/v1/sessions",
            json={"title": "Compacted", "context_policy": "summary_and_last_turn"},
        )
        assert create_response.status_code == 200
        session_id = create_response.json()["session_id"]
        assert create_response.json()["context_policy"] == "summary_and_last_turn"

        first_chat = client.post(
            "/v1/chat/completions",
            json={
                "model": gguf_model_id,
                "session_id": session_id,
                "messages": [{"role": "user", "content": "First compacted note"}],
            },
        )
        second_chat = client.post(
            "/v1/chat/completions",
            json={
                "model": gguf_model_id,
                "session_id": session_id,
                "messages": [{"role": "user", "content": "Second compacted note"}],
            },
        )
        third_chat = client.post(
            "/v1/chat/completions",
            json={
                "model": gguf_model_id,
                "session_id": session_id,
                "messages": [{"role": "user", "content": "Third compacted note"}],
                "include_prompt_trace": True,
            },
        )

    assert first_chat.status_code == 200
    assert second_chat.status_code == 200
    assert third_chat.status_code == 200
    assert third_chat.json()["prompt_trace"]["message_count"] == 4
    assert third_chat.json()["prompt_trace"]["message_roles"] == ["system", "user", "assistant", "user"]


def test_streaming_chat_completes_when_session_persistence_fails(
    session_enabled_settings,
    services_with_fake_runtime_session_enabled,
    monkeypatch,
    caplog,
) -> None:
    app = create_app(session_enabled_settings, services=services_with_fake_runtime_session_enabled)

    def fail_record_turn(*args, **kwargs):
        raise RuntimeError("session store unavailable")

    monkeypatch.setattr(services_with_fake_runtime_session_enabled.session_history_service, "record_turn", fail_record_turn)

    with TestClient(app) as client:
        scan_response = client.post("/v1/models/scan", json={})
        manifests = scan_response.json()["manifests"]
        gguf_model_id = next(manifest["model_id"] for manifest in manifests if manifest["format_type"] == "gguf")

        create_response = client.post("/v1/sessions", json={"title": "stream failure"})
        session_id = create_response.json()["session_id"]

        with client.stream(
            "POST",
            "/v1/chat/completions",
            json={
                "model": gguf_model_id,
                "session_id": session_id,
                "messages": [{"role": "user", "content": "Stream this despite persistence issues"}],
                "stream": True,
            },
        ) as stream_response:
            assert stream_response.status_code == 200
            data_lines = [
                line.removeprefix("data: ")
                for line in stream_response.iter_lines()
                if line and line.startswith("data: ")
            ]

        session_detail = client.get(f"/v1/sessions/{session_id}").json()

    assert data_lines[-1] == "[DONE]"
    stream_chunks = [json.loads(line) for line in data_lines[:-1]]
    assert stream_chunks[-1]["metadata"]["model"]["resolved_model_id"] == gguf_model_id
    assert stream_chunks[-1]["metadata"]["timing"]["total_milliseconds"] >= 0
    streamed_text = "".join(
        chunk["choices"][0]["delta"].get("content", "")
        for chunk in stream_chunks
        if chunk["choices"][0]["delta"].get("content")
    )
    assert streamed_text == "Echo: Stream this despite persistence issues"
    assert session_detail["turn_count"] == 0
    assert "Streaming chat session completion callback failed." in caplog.text


def test_session_endpoints_reject_persistence_when_privacy_mode_enabled(app_with_fake_runtime) -> None:
    with TestClient(app_with_fake_runtime) as client:
        create_response = client.post("/v1/sessions", json={"title": "Blocked"})
        chat_response = client.post(
            "/v1/chat/completions",
            json={
                "session_id": "existing-session",
                "messages": [{"role": "user", "content": "Persist this"}],
            },
        )

    assert create_response.status_code == 403
    assert create_response.json()["error"]["code"] == "privacy_mode_enabled"
    assert chat_response.status_code == 403
    assert chat_response.json()["error"]["code"] == "privacy_mode_enabled"


def test_chat_and_responses_support_prompt_override_traces(
    app_with_fake_runtime,
    sample_prompt_assets,
) -> None:
    with TestClient(app_with_fake_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        manifests = scan_response.json()["manifests"]
        gguf_model_id = next(manifest["model_id"] for manifest in manifests if manifest["format_type"] == "gguf")

        chat_response = client.post(
            "/v1/chat/completions",
            json={
                "model": gguf_model_id,
                "messages": [{"role": "user", "content": "Summarize milestone eight"}],
                "developer_prompt": "Keep it terse.",
                "pretext_path": str(sample_prompt_assets["pretext"]),
                "skills_path": str(sample_prompt_assets["skill"]),
                "response_format_path": str(sample_prompt_assets["response_format"]),
                "tools_path": str(sample_prompt_assets["tools"]),
                "mcp_tools_path": str(sample_prompt_assets["mcp_tools"]),
                "include_prompt_trace": True,
            },
        )
        responses_api = client.post(
            "/v1/responses",
            json={
                "model": gguf_model_id,
                "input": "Turn this into a status object",
                "developer_prompt": "Return only the essentials.",
                "pretext_path": str(sample_prompt_assets["pretext"]),
                "output_schema_path": str(sample_prompt_assets["output_schema"]),
                "include_prompt_trace": True,
            },
        )

    assert chat_response.status_code == 200
    assert chat_response.json()["metadata"]["model"]["resolved_model_id"] == gguf_model_id
    assert chat_response.json()["metadata"]["routing"]["kind"] == "model_router"
    assert chat_response.json()["metadata"]["timing"]["load_milliseconds"] >= 0
    chat_trace = chat_response.json()["prompt_trace"]
    assert chat_trace["selected_template"] == "tool_structured_output"
    assert chat_trace["model_prompt_template"]["id"] == "llama-instruct-v1"
    assert "[INST]" in chat_trace["serialized_model_prompt"]
    assert [override["source"] for override in chat_trace["overrides"]] == [
        "pretext_file",
        "developer_prompt",
        "skills_file",
        "tools_file",
        "mcp_tools_file",
        "response_format_file",
    ]
    assert chat_trace["output_contract"]["format"] == "json_schema"
    assert any(tool["name"] == "local_lookup" for tool in chat_trace["tool_plan"])
    mcp_tool = next(tool for tool in chat_trace["tool_plan"] if tool["name"] == "search_milestones")
    assert mcp_tool["source"] == "mcp_tools_file"
    assert mcp_tool["mcp_server"] == "roadmap"
    assert mcp_tool["metadata_trusted"] is False
    chat_structured_output = chat_response.json()["structured_output"]
    assert chat_structured_output["enforcement"] == "decode_time"
    assert chat_structured_output["decoder_enforced"] is True
    assert chat_structured_output["fallback_used"] is False
    assert chat_structured_output["validation"]["state"] == "valid"
    assert chat_structured_output["parsed_output"] == {"summary": "ok", "status": "ok"}

    assert responses_api.status_code == 200
    assert responses_api.json()["metadata"]["model"]["resolved_model_id"] == gguf_model_id
    assert responses_api.json()["metadata"]["routing"]["kind"] == "model_router"
    assert responses_api.json()["metadata"]["timing"]["execute_milliseconds"] >= 0
    response_trace = responses_api.json()["prompt_trace"]
    assert response_trace["selected_template"] == "structured_output"
    assert response_trace["model_prompt_template"]["id"] == "llama-instruct-v1"
    assert "[INST]" in response_trace["serialized_model_prompt"]
    assert [override["source"] for override in response_trace["overrides"]] == [
        "pretext_file",
        "developer_prompt",
        "output_schema_file",
    ]
    responses_structured_output = responses_api.json()["structured_output"]
    assert responses_structured_output["enforcement"] == "decode_time"
    assert responses_structured_output["decoder_enforced"] is True
    assert responses_structured_output["fallback_used"] is False
    assert responses_structured_output["validation"]["state"] == "valid"
    assert responses_structured_output["parsed_output"] == {"summary": "ok", "status": "ok"}


def test_chat_and_responses_accept_response_format_contracts(app_with_fake_runtime) -> None:
    with TestClient(app_with_fake_runtime) as client:
        scan_response = client.post("/v1/models/scan", json={})
        manifests = scan_response.json()["manifests"]
        gguf_model_id = next(manifest["model_id"] for manifest in manifests if manifest["format_type"] == "gguf")

        chat_response = client.post(
            "/v1/chat/completions",
            json={
                "model": gguf_model_id,
                "messages": [{"role": "user", "content": "Return ok"}],
                "response_format": {
                    "type": "grammar",
                    "name": "literal_ok",
                    "syntax": "ebnf",
                    "grammar": 'root ::= "ok"',
                },
                "include_prompt_trace": True,
            },
        )
        responses_api = client.post(
            "/v1/responses",
            json={
                "model": gguf_model_id,
                "input": "Return a structured summary",
                "response_format": {
                    "type": "json_schema",
                    "name": "status_summary",
                    "schema": {
                        "type": "object",
                        "properties": {"summary": {"type": "string"}},
                        "required": ["summary"],
                        "additionalProperties": False,
                    },
                },
                "include_prompt_trace": True,
            },
        )

    assert chat_response.status_code == 200
    assert chat_response.json()["prompt_trace"]["output_contract"]["format"] == "grammar"
    chat_structured_output = chat_response.json()["structured_output"]
    assert chat_structured_output["contract"]["type"] == "grammar"
    assert chat_structured_output["enforcement"] == "decode_time"
    assert chat_structured_output["decoder_enforced"] is True
    assert chat_structured_output["fallback_used"] is False
    assert chat_structured_output["validation"]["state"] == "valid"
    assert chat_structured_output["validation"]["validator"] == "grammar"

    assert responses_api.status_code == 200
    assert responses_api.json()["prompt_trace"]["overrides"][-1]["source"] == "response_format"
    responses_structured_output = responses_api.json()["structured_output"]
    assert responses_structured_output["contract"]["type"] == "json_schema"
    assert responses_structured_output["enforcement"] == "decode_time"
    assert responses_structured_output["decoder_enforced"] is True
    assert responses_structured_output["fallback_used"] is False
    assert responses_structured_output["validation"]["state"] == "valid"
    assert responses_structured_output["parsed_output"] == {"summary": "ok"}
